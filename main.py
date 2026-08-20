import os
import io
import re
import json
import time
import uuid
import logging
import datetime as dt
from decimal import Decimal
from typing import Optional, List, Union, Tuple

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field

from sqlalchemy import (
    create_engine, Column, Integer, String, Float, Boolean, DateTime,
    ForeignKey, Text, Enum as SAEnum, func
)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, Session, joinedload
from sqlalchemy.exc import IntegrityError

from passlib.context import CryptContext
from jose import jwt, JWTError

import boto3
from botocore.exceptions import ClientError

# PDF report generation (pip install reportlab)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage

logger = logging.getLogger("ktc_ipfms")
logging.basicConfig(level=logging.INFO)

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

DATABASE_URL = os.getenv("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    # Render / Heroku style URLs use the old "postgres://" scheme; SQLAlchemy
    # (via psycopg2) needs "postgresql://".
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

if not DATABASE_URL:
    DATABASE_URL = "sqlite:///./ktc.db"

JWT_SECRET = os.getenv("JWT_SECRET", "change-this-secret-in-production")
JWT_ALGORITHM = "HS256"
# The JWT's own lifetime is intentionally long — actual session length is now
# enforced client-side by a 10-minute *inactivity* timer (see the frontend's
# idle-logout logic) instead of by the token itself expiring on a fixed
# schedule. A fixed 10-minute token previously logged active users out mid-task
# regardless of whether they were still using the system.
ACCESS_TOKEN_EXPIRE_MINUTES = 480

# --------------------------------------------------------------------------
# Backblaze B2 (S3-compatible) object storage
# --------------------------------------------------------------------------
B2_BUCKET_NAME = os.getenv("B2_BUCKET_NAME", "uploads-dir")
B2_ENDPOINT_URL = os.getenv("B2_ENDPOINT_URL", "https://s3.us-east-005.backblazeb2.com")
B2_KEY_ID = os.getenv("B2_KEY_ID", "0055ca7845641d30000000002")
B2_APPLICATION_KEY = os.getenv("B2_APPLICATION_KEY", "K005NNeGM9r28ujQ3jvNEQy2zUiu0TI")
B2_DOCUMENTS_FOLDER = "ktc-documents"
B2_SIGNATURES_FOLDER = "ktc-signatures"

b2_client = boto3.client(
    "s3",
    endpoint_url=B2_ENDPOINT_URL,
    aws_access_key_id=B2_KEY_ID,
    aws_secret_access_key=B2_APPLICATION_KEY,
)

CORS_ORIGINS = os.getenv("CORS_ORIGINS", "*")
origins = ["*"] if CORS_ORIGINS.strip() == "*" else [o.strip() for o in CORS_ORIGINS.split(",")]

connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, connect_args=connect_args, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

ROLES = [
    "admin",       # System Administrator
    "staff",       # Staff Member
    "hod",         # Head of Department
    "treasurer",   # Senior Treasurer
    "clerk",       # Town Clerk
    "auditor",     # Internal Auditor
]

# --------------------------------------------------------------------------
# Lightweight in-memory response cache
# --------------------------------------------------------------------------
# The Work Plan & Budget table was slow to render because every request
# recomputed committed/available balances with a fresh DB round-trip PER
# ROW (see the old BudgetCode.committed_amount property, which opened a new
# SessionLocal() session per budget code — an N+1 query pattern). Two things
# fix this:
#   1. list_budget_codes now computes committed amounts for the whole page
#      in a single grouped query (see _bulk_committed_amounts) instead of
#      one query per row.
#   2. The fully-serialised response for a given (work_plan_id,
#      department_id, search) combination is cached in memory for a short
#      TTL, so repeat requests (e.g. re-opening the Work Plan & Budget tab,
#      or switching between filters that were already viewed) are served
#      instantly without touching the database at all.
# The cache is invalidated proactively whenever the underlying data changes
# (budget code create/update/import, requisition submit/approve/reject,
# department or work plan creation), so results never go stale beyond
# that TTL by more than the time it takes those write paths to run.
_CACHE_TTL_SECONDS = 90
_response_cache: dict = {}


def _cache_get(key: str):
    entry = _response_cache.get(key)
    if not entry:
        return None
    ts, value = entry
    if time.time() - ts > _CACHE_TTL_SECONDS:
        _response_cache.pop(key, None)
        return None
    return value


def _cache_set(key: str, value):
    _response_cache[key] = (time.time(), value)


def _cache_invalidate_prefix(prefix: str):
    for k in [k for k in _response_cache if k.startswith(prefix)]:
        _response_cache.pop(k, None)


def _invalidate_budget_code_caches():
    _cache_invalidate_prefix("budget_codes:")
    _cache_invalidate_prefix("dashboard_stats:")


def _invalidate_revenue_source_caches():
    _cache_invalidate_prefix("revenue_sources:")


# --------------------------------------------------------------------------
# Shared numeric amount parser
# --------------------------------------------------------------------------
_THOUSANDS_SPACE_RE = re.compile(r"(?<=\d)[\s\u00A0\u2009\u202F](?=\d)")
_AMOUNT_CURRENCY_NOISE = ("UGX", "Ugx", "ugx", "USH", "Ush", "ush", "/=", "=", "%")


def parse_amount_verbose(v):
    """Parse any raw value (native number, text, None) into a float.

    Returns (value, ok, original_text):
      - value: the parsed float (0.0 if v was empty/None, or if parsing
        ultimately failed)
      - ok: False only when a genuinely non-empty, non-numeric value had to
        be discarded — lets callers that care (the Excel importer) warn the
        user instead of staying silent
      - original_text: the original value as text, for error messages
    """
    if v is None:
        return 0.0, True, ""
    if isinstance(v, bool):
        return 0.0, True, str(v)
    if isinstance(v, (int, float, Decimal)):
        return float(v), True, str(v)

    original = str(v)
    s = original.strip()
    if s == "":
        return 0.0, True, s

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    elif s.startswith("-"):
        negative = True
        s = s[1:].strip()

    s = s.replace("\u00A0", " ").replace("\u2009", " ").replace("\u202F", " ")

    for token in _AMOUNT_CURRENCY_NOISE:
        s = s.replace(token, "")
    s = s.strip()

    s = s.replace(",", "")
    s = _THOUSANDS_SPACE_RE.sub("", s)
    s = s.strip()

    if s == "" or s == "-":
        return 0.0, True, original.strip()

    try:
        result = float(s)
    except (TypeError, ValueError):
        return 0.0, False, original.strip()

    return (-result if negative else result), True, original.strip()


def parse_amount(v) -> float:
    """Best-effort float conversion — never raises, defaults to 0.0."""
    value, _ok, _original = parse_amount_verbose(v)
    return value


_LEADING_NUMBER_RE = re.compile(r"-?\d[\d,]*(?:\.\d+)?")


def parse_leading_number(v) -> Tuple[float, bool]:
    """For fields (Baseline Value / Planned Target) that are frequently
    narrative or multi-part text rather than a single number — e.g.
    "25% of Council area mapped for vectors" or a per-quarter list like
    "12; 4; 4; 4; 4; 4; 4" — pull out the first number found anywhere in
    the text so the value is still usable in numeric aggregates/charts,
    rather than collapsing straight to 0.

    Returns (value, was_full_text): was_full_text is True when the raw
    value wasn't a clean standalone number (i.e. there was extra text
    around/after the number, or no number at all) — signals to the caller
    that the original text is worth preserving in a *_note column.
    """
    value, ok, original = parse_amount_verbose(v)
    if ok:
        # Already a clean number (or genuinely empty) — nothing extra to
        # preserve.
        return value, False
    match = _LEADING_NUMBER_RE.search(original)
    if not match:
        return 0.0, True
    try:
        return float(match.group(0).replace(",", "")), True
    except ValueError:
        return 0.0, True


# --------------------------------------------------------------------------
# Models
# --------------------------------------------------------------------------

class Department(Base):
    __tablename__ = "departments"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(150), unique=True, nullable=False)
    code = Column(String(20), unique=True, nullable=False)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    users = relationship("User", back_populates="department")
    budget_codes = relationship("BudgetCode", back_populates="department")


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    full_name = Column(String(150), nullable=False)
    email = Column(String(150), unique=True, nullable=False, index=True)
    hashed_password = Column(String(255), nullable=False)
    role = Column(String(20), nullable=False)
    department_id = Column(Integer, ForeignKey("departments.id"), nullable=True)
    position = Column(String(150), nullable=True)
    telephone = Column(String(40), nullable=True)
    signature_path = Column(String(500), nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    department = relationship("Department", back_populates="users")

    @property
    def signature_url(self):
        return f"/files/{self.signature_path}" if self.signature_path else None


class WorkPlan(Base):
    __tablename__ = "work_plans"
    id = Column(Integer, primary_key=True, index=True)
    financial_year = Column(String(20), nullable=False)   # e.g. "2026/27"
    title = Column(String(200), nullable=False)
    # Headings shown above the 4 Work Plan & Budget tables for this work
    # plan. Maintained here (via the New/Edit Work Plan modal) rather than
    # per-table, so switching the work plan dropdown switches every table's
    # heading in one go, and there is exactly one place to edit them.
    title_revenue_summary = Column(String(500), nullable=True)
    title_dept_summary = Column(String(500), nullable=True)
    title_revenue_detail = Column(String(500), nullable=True)
    title_main_table = Column(String(500), nullable=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    budget_codes = relationship("BudgetCode", back_populates="work_plan")
    revenue_sources = relationship("RevenueSource", back_populates="work_plan")


class BudgetCode(Base):
    __tablename__ = "budget_codes"
    id = Column(Integer, primary_key=True, index=True)
    work_plan_id = Column(Integer, ForeignKey("work_plans.id"), nullable=False)
    department_id = Column(Integer, ForeignKey("departments.id"), nullable=False)
    # These are free-text narrative fields copied verbatim from Council
    # work-plan workbooks, which routinely run well past 255 characters
    # (e.g. a "Budget Output Description" of 400+ characters is normal for
    # this dataset) — stored as Text (unbounded) rather than a bounded
    # VARCHAR so a long but perfectly legitimate description never causes
    # the whole import batch to fail with a Postgres
    # StringDataRightTruncation error.
    service_area = Column(Text)
    code = Column(String(30), nullable=False)                  # Budget Output Code
    output_description = Column(Text, nullable=False)          # Budget Output Description
    programme = Column(Text)
    sub_programme = Column(Text)
    piap_output_description = Column(Text)
    piap_output_indicator = Column(Text)
    unit_of_measure = Column(String(100))
    baseline_value = Column(Float, default=0)
    # Full original text for Baseline Value / Planned Target as entered in
    # the source workbook. Many PIAP indicators record these as narrative
    # or multi-part values (e.g. "25% of Council area mapped for vectors",
    # "12; 4; 4; 4; 4; 4; 4" for per-quarter figures) rather than a single
    # number. baseline_value/planned_target hold a best-effort numeric
    # figure (a leading number extracted from the text, or 0) so existing
    # charts/aggregates keep working; these _note columns preserve the
    # full original text so nothing is lost on import.
    baseline_note = Column(Text)
    planned_target = Column(Float, default=0)
    target_note = Column(Text)
    actual_output = Column(Text)
    q1_amount = Column(Float, default=0)
    q2_amount = Column(Float, default=0)
    q3_amount = Column(Float, default=0)
    q4_amount = Column(Float, default=0)
    funding_source = Column(String(255), default="Local Revenue")  # Revenue Source
    # Often a multi-line list of several named officers/committees in
    # practice (e.g. "Town Mayor, Town Clerk, LCII Chairpersons, Ward
    # Development Committees, Clerk to Council") — Text avoids truncation.
    responsible_party = Column(Text)

    work_plan = relationship("WorkPlan", back_populates="budget_codes")
    department = relationship("Department", back_populates="budget_codes")
    activities = relationship("Activity", back_populates="budget_code")

    @property
    def allocated_amount(self):
        return (
            parse_amount(self.q1_amount) + parse_amount(self.q2_amount) +
            parse_amount(self.q3_amount) + parse_amount(self.q4_amount)
        )

    @property
    def committed_amount(self):
        # NOTE: kept for any callers that need a single budget code's
        # committed amount in isolation. Bulk endpoints (list_budget_codes,
        # dashboard_stats) should use _bulk_committed_amounts() instead to
        # avoid the N+1 query pattern this property has when used per-row
        # in a loop.
        db = SessionLocal()
        try:
            reqs = db.query(Requisition).filter(
                Requisition.budget_code_id == self.id,
                Requisition.status.notin_(["rejected", "returned", "draft"])
            ).all()
            return sum(r.amount_requested for r in reqs)
        finally:
            db.close()

    @property
    def available_balance(self):
        return self.allocated_amount - self.committed_amount


class RevenueSource(Base):
    __tablename__ = "revenue_sources"
    id = Column(Integer, primary_key=True, index=True)
    work_plan_id = Column(Integer, ForeignKey("work_plans.id"), nullable=False)
    pbs_fund_code = Column(String(50))
    source_of_financing_name = Column(String(255), nullable=False)
    functional_definition = Column(Text)
    # Manually-entered Approved Budget Amount. This is now used ONLY as a
    # fallback for revenue sources that have no sub rows (revenue items) —
    # once at least one RevenueSourceItem exists under this source, the
    # Approved Budget Amount / Category Total is obtained automatically as
    # the sum of those sub rows instead (see revenue_source_to_out below),
    # matching the "Revenue Entry Table" behaviour in the UI.
    approved_budget_amount = Column(Float, default=0)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    work_plan = relationship("WorkPlan", back_populates="revenue_sources")
    items = relationship(
        "RevenueSourceItem",
        back_populates="revenue_source",
        cascade="all, delete-orphan",
        order_by="RevenueSourceItem.id",
    )


class RevenueSourceItem(Base):
    """A single sub row (revenue item) under a RevenueSource category, e.g.
    under PBS Fund Code / Source of Financing "Central Government Transfers
    (GoU)" there might be sub rows "Unconditional Grant – Wage",
    "Unconditional Grant – Non-Wage", etc., each with its own Approved
    Estimate. The parent RevenueSource's Category Total (and therefore the
    Approved Budget Amount shown in the Summary of Sources of Revenue) is
    the sum of these sub rows whenever any exist."""
    __tablename__ = "revenue_source_items"
    id = Column(Integer, primary_key=True, index=True)
    revenue_source_id = Column(Integer, ForeignKey("revenue_sources.id"), nullable=False)
    description = Column(Text, nullable=False)
    amount = Column(Float, default=0)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    revenue_source = relationship("RevenueSource", back_populates="items")


class Activity(Base):
    __tablename__ = "activities"
    id = Column(Integer, primary_key=True, index=True)
    budget_code_id = Column(Integer, ForeignKey("budget_codes.id"), nullable=False)
    name = Column(String(255), nullable=False)
    quarter = Column(String(10), default="Q1")  # Q1-Q4
    is_active = Column(Boolean, default=True)

    budget_code = relationship("BudgetCode", back_populates="activities")


class Requisition(Base):
    __tablename__ = "requisitions"
    id = Column(Integer, primary_key=True, index=True)
    ref_no = Column(String(60), unique=True, nullable=False)
    requester_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    department_id = Column(Integer, ForeignKey("departments.id"), nullable=False)
    budget_code_id = Column(Integer, ForeignKey("budget_codes.id"), nullable=True)
    activity_id = Column(Integer, ForeignKey("activities.id"), nullable=True)
    activity_details = Column(Text)
    subject = Column(String(255), nullable=True)
    # Financial Year / Quarter this requisition is being drawn against, and
    # the requisitioner's typed-in name/position on the paper-form replica
    # (kept separate from the User account, since the Requisitioner field is
    # now free text the person fills in themselves rather than a read-only
    # mirror of their account name).
    financial_year = Column(String(20), nullable=True)
    quarter = Column(String(10), nullable=True)
    requester_name = Column(String(150), nullable=True)
    requester_position = Column(String(150), nullable=True)
    # Requisitioner's mobile/telephone number, typed on the paper-form
    # replica's Requisitioner row (kept separate from the User account's
    # own telephone, same reasoning as requester_name/requester_position).
    requester_mobile = Column(String(40), nullable=True)
    # Free-typed Budget Output Code as entered on the form. Kept alongside
    # budget_code_id (which links to a real BudgetCode record when the
    # typed code matches one on file) so a code that doesn't match any
    # known BudgetCode is still captured rather than silently dropped.
    budget_output_code_text = Column(String(50), nullable=True)
    payment_voucher_number = Column(String(100), nullable=True)
    line_items = Column(Text, nullable=True)
    # Free-form JSON blob holding the fields specific to the Cheque Payment
    # Voucher section of the combined modal (Dr. To, Cheque No, Address,
    # Authority, Approved Vote, Account No, ledger folio, etc). Kept as one
    # column rather than a dozen new ones since these fields are entered
    # together and displayed together.
    voucher_data = Column(Text, nullable=True)
    amount_requested = Column(Float, nullable=False)
    status = Column(String(20), default="draft")
    current_stage = Column(String(20), default="hod")  # hod / treasurer / clerk / done
    created_at = Column(DateTime, default=dt.datetime.utcnow)
    updated_at = Column(DateTime, default=dt.datetime.utcnow)

    requester = relationship("User", foreign_keys=[requester_id])
    department = relationship("Department")
    budget_code = relationship("BudgetCode")
    activity = relationship("Activity")
    approvals = relationship("ApprovalHistory", back_populates="requisition", order_by="ApprovalHistory.id")
    documents = relationship("Document", back_populates="requisition")
    accountability = relationship("AccountabilityRecord", back_populates="requisition", uselist=False)


class ApprovalHistory(Base):
    __tablename__ = "approval_history"
    id = Column(Integer, primary_key=True, index=True)
    requisition_id = Column(Integer, ForeignKey("requisitions.id"), nullable=False)
    stage = Column(String(20), nullable=False)   # hod / treasurer / clerk
    actor_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    action = Column(String(20), nullable=False)  # approve / reject / return
    comments = Column(Text)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    requisition = relationship("Requisition", back_populates="approvals")
    actor = relationship("User")


class AccountabilityRecord(Base):
    __tablename__ = "accountability_records"
    id = Column(Integer, primary_key=True, index=True)
    requisition_id = Column(Integer, ForeignKey("requisitions.id"), unique=True, nullable=False)
    auditor_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    status = Column(String(20), default="pending")  # pending / verified / flagged
    remarks = Column(Text)
    updated_at = Column(DateTime, default=dt.datetime.utcnow)

    requisition = relationship("Requisition", back_populates="accountability")
    auditor = relationship("User")


class Document(Base):
    __tablename__ = "documents"
    id = Column(Integer, primary_key=True, index=True)
    requisition_id = Column(Integer, ForeignKey("requisitions.id"), nullable=False)
    filename = Column(String(255), nullable=False)
    stored_path = Column(String(500), nullable=False)
    doc_type = Column(String(50), default="supporting")  # supporting / receipt / voucher / attendance / other
    uploaded_by = Column(Integer, ForeignKey("users.id"), nullable=False)
    created_at = Column(DateTime, default=dt.datetime.utcnow)

    requisition = relationship("Requisition", back_populates="documents")


class Notification(Base):
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    message = Column(String(500), nullable=False)
    category = Column(String(30), default="info")
    is_read = Column(Boolean, default=False)
    created_at = Column(DateTime, default=dt.datetime.utcnow)
    link_requisition_id = Column(Integer, ForeignKey("requisitions.id"), nullable=True)


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    action = Column(String(150), nullable=False)
    details = Column(Text)
    created_at = Column(DateTime, default=dt.datetime.utcnow)


Base.metadata.create_all(bind=engine)


def _run_lightweight_migrations():
    statements = [
        "ALTER TABLE budget_codes ADD COLUMN indicator VARCHAR(255)",
        "ALTER TABLE budget_codes ADD COLUMN q1_amount FLOAT DEFAULT 0",
        "ALTER TABLE budget_codes ADD COLUMN q2_amount FLOAT DEFAULT 0",
        "ALTER TABLE budget_codes ADD COLUMN q3_amount FLOAT DEFAULT 0",
        "ALTER TABLE budget_codes ADD COLUMN q4_amount FLOAT DEFAULT 0",
        "ALTER TABLE budget_codes ADD COLUMN service_area VARCHAR(150)",
        "ALTER TABLE budget_codes ADD COLUMN piap_output_description VARCHAR(255)",
        "ALTER TABLE budget_codes ADD COLUMN piap_output_indicator VARCHAR(255)",
        "ALTER TABLE budget_codes ADD COLUMN actual_output VARCHAR(255)",
        "ALTER TABLE budget_codes ADD COLUMN responsible_party VARCHAR(150)",
        # Preserves the full original Baseline Value / Planned Target text
        # for rows where it's narrative or multi-part (e.g. "25% of
        # Council area mapped for vectors") rather than a clean number —
        # baseline_value/planned_target keep a best-effort numeric figure
        # for existing charts/aggregates, these columns hold the source
        # text so nothing is lost on import.
        "ALTER TABLE budget_codes ADD COLUMN baseline_note TEXT",
        "ALTER TABLE budget_codes ADD COLUMN target_note TEXT",
        "ALTER TABLE users ADD COLUMN position VARCHAR(150)",
        "ALTER TABLE users ADD COLUMN telephone VARCHAR(40)",
        "ALTER TABLE users ADD COLUMN signature_path VARCHAR(500)",
        "ALTER TABLE requisitions ADD COLUMN subject VARCHAR(255)",
        "ALTER TABLE requisitions ADD COLUMN line_items TEXT",
        "ALTER TABLE requisitions ADD COLUMN payment_voucher_number VARCHAR(100)",
        "ALTER TABLE requisitions ADD COLUMN voucher_data TEXT",
        # Budget Code / Activity are no longer filled in on the requisition
        # entry screen (removed per Council request) — a requisition can now
        # be saved without one, so the column must accept NULL. Postgres-only
        # syntax; fails harmlessly on SQLite, which never enforced this.
        "ALTER TABLE requisitions ALTER COLUMN budget_code_id DROP NOT NULL",
        # ref_no used to be capped at 40 chars ("KTC-REQ-YYYY-00001"); the new
        # "KTC-RQ-YY-MM-DD-<payment voucher number>" format can run longer
        # depending on what the PV number looks like, so widen the column.
        "ALTER TABLE requisitions ALTER COLUMN ref_no TYPE VARCHAR(60)",
        # Widen narrative BudgetCode columns on an already-deployed Postgres
        # database from bounded VARCHAR to unbounded TEXT, so long but
        # legitimate descriptions from imported work-plan workbooks (e.g.
        # 400+ character Budget Output Descriptions) stop being rejected
        # with a StringDataRightTruncation error. These are Postgres-only
        # syntax and simply fail harmlessly (caught below) against SQLite,
        # which doesn't enforce VARCHAR length limits in the first place.
        "ALTER TABLE budget_codes ALTER COLUMN service_area TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN output_description TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN programme TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN sub_programme TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN piap_output_description TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN piap_output_indicator TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN actual_output TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN responsible_party TYPE TEXT",
        "ALTER TABLE budget_codes ALTER COLUMN unit_of_measure TYPE VARCHAR(100)",
        "ALTER TABLE budget_codes ALTER COLUMN funding_source TYPE VARCHAR(255)",
        # Per-work-plan table headings — each work plan now carries its own
        # set of the 4 table titles shown on the Work Plan & Budget view, so
        # switching the "Annual Work Plan" dropdown switches the headings
        # too, and editing/creating a work plan is the only place these are
        # maintained (no separate title-management UI on each table).
        "ALTER TABLE work_plans ADD COLUMN title_revenue_summary VARCHAR(500)",
        "ALTER TABLE work_plans ADD COLUMN title_dept_summary VARCHAR(500)",
        "ALTER TABLE work_plans ADD COLUMN title_revenue_detail VARCHAR(500)",
        "ALTER TABLE work_plans ADD COLUMN title_main_table VARCHAR(500)",
        # Financial Year / Quarter drawn against, and the requisitioner's
        # typed-in name/position on the Funds Requisition Form replica.
        "ALTER TABLE requisitions ADD COLUMN financial_year VARCHAR(20)",
        "ALTER TABLE requisitions ADD COLUMN quarter VARCHAR(10)",
        "ALTER TABLE requisitions ADD COLUMN requester_name VARCHAR(150)",
        "ALTER TABLE requisitions ADD COLUMN requester_position VARCHAR(150)",
        # Requisitioner's Mob. No. (telephone) and the free-typed Budget
        # Output Code shown on the Funds Requisition Form replica.
        "ALTER TABLE requisitions ADD COLUMN requester_mobile VARCHAR(40)",
        "ALTER TABLE requisitions ADD COLUMN budget_output_code_text VARCHAR(50)",
    ]
    with engine.connect() as conn:
        for stmt in statements:
            try:
                conn.exec_driver_sql(stmt)
                conn.commit()
            except Exception:
                conn.rollback()  # column already exists (or backend quirk) — ignore

    try:
        with engine.connect() as conn:
            conn.exec_driver_sql(
                "UPDATE budget_codes SET piap_output_indicator = indicator "
                "WHERE (piap_output_indicator IS NULL OR piap_output_indicator = '') "
                "AND indicator IS NOT NULL AND indicator <> ''"
            )
            conn.commit()
    except Exception:
        pass


_run_lightweight_migrations()

# --------------------------------------------------------------------------
# Schemas
# --------------------------------------------------------------------------

class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"
    role: str
    full_name: str
    user_id: int


class LoginIn(BaseModel):
    email: EmailStr
    password: str
    role: Optional[str] = None


class UserOut(BaseModel):
    id: int
    full_name: str
    email: str
    role: str
    department_id: Optional[int] = None
    position: Optional[str] = None
    telephone: Optional[str] = None
    is_active: bool
    signature_url: Optional[str] = None

    class Config:
        from_attributes = True


class UserCreate(BaseModel):
    full_name: str
    email: EmailStr
    password: str
    role: str
    department_id: Optional[int] = None
    position: Optional[str] = None
    telephone: Optional[str] = None


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    role: Optional[str] = None
    department_id: Optional[int] = None
    position: Optional[str] = None
    telephone: Optional[str] = None


class DepartmentIn(BaseModel):
    name: str
    code: str


class DepartmentOut(DepartmentIn):
    id: int
    class Config:
        from_attributes = True


class WorkPlanIn(BaseModel):
    financial_year: str
    title: str
    title_revenue_summary: Optional[str] = None
    title_dept_summary: Optional[str] = None
    title_revenue_detail: Optional[str] = None
    title_main_table: Optional[str] = None


class WorkPlanOut(WorkPlanIn):
    id: int
    is_active: bool
    class Config:
        from_attributes = True


class BudgetCodeIn(BaseModel):
    work_plan_id: int
    department_id: int
    service_area: Optional[str] = None
    code: str
    output_description: str
    programme: Optional[str] = None
    sub_programme: Optional[str] = None
    piap_output_description: Optional[str] = None
    piap_output_indicator: Optional[str] = None
    unit_of_measure: Optional[str] = None
    baseline_value: Union[float, int, str] = 0
    baseline_note: Optional[str] = None
    planned_target: Union[float, int, str] = 0
    target_note: Optional[str] = None
    actual_output: Optional[str] = None
    q1_amount: Union[float, int, str] = 0
    q2_amount: Union[float, int, str] = 0
    q3_amount: Union[float, int, str] = 0
    q4_amount: Union[float, int, str] = 0
    funding_source: str = "Local Revenue"
    responsible_party: Optional[str] = None


class BudgetCodeUpdate(BaseModel):
    service_area: Optional[str] = None
    code: Optional[str] = None
    output_description: Optional[str] = None
    programme: Optional[str] = None
    sub_programme: Optional[str] = None
    piap_output_description: Optional[str] = None
    piap_output_indicator: Optional[str] = None
    unit_of_measure: Optional[str] = None
    baseline_value: Optional[Union[float, int, str]] = None
    baseline_note: Optional[str] = None
    planned_target: Optional[Union[float, int, str]] = None
    target_note: Optional[str] = None
    actual_output: Optional[str] = None
    q1_amount: Optional[Union[float, int, str]] = None
    q2_amount: Optional[Union[float, int, str]] = None
    q3_amount: Optional[Union[float, int, str]] = None
    q4_amount: Optional[Union[float, int, str]] = None
    funding_source: Optional[str] = None
    responsible_party: Optional[str] = None


class BudgetCodeOut(BaseModel):
    id: int
    work_plan_id: int
    department_id: int
    department_name: Optional[str] = None
    # Only populated/used for the Annual Work Plan & Budget Estimates table's
    # Department column (per Council convention of showing the PBS code
    # alongside the name there, e.g. "090: Community Based Services").
    # Every other screen in the system shows department_name alone — see
    # _dept_label()'s docstring for the strict-separation rule.
    department_code_and_name: Optional[str] = None
    service_area: Optional[str] = None
    code: str
    output_description: str
    programme: Optional[str] = None
    sub_programme: Optional[str] = None
    piap_output_description: Optional[str] = None
    piap_output_indicator: Optional[str] = None
    unit_of_measure: Optional[str] = None
    baseline_value: float
    baseline_note: Optional[str] = None
    planned_target: float
    target_note: Optional[str] = None
    actual_output: Optional[str] = None
    q1_amount: float
    q2_amount: float
    q3_amount: float
    q4_amount: float
    funding_source: str
    responsible_party: Optional[str] = None
    allocated_amount: float
    committed_amount: float
    available_balance: float

    class Config:
        from_attributes = True


class BudgetCodeImportResult(BaseModel):
    created: int
    skipped: int
    departments_created: int = 0
    errors: List[str] = []
    # Total warning count before truncation, so the UI can show "showing
    # 30 of N" instead of silently hiding warnings past the first 30.
    total_warnings: int = 0


class BudgetCodeClearResult(BaseModel):
    deleted: int
    skipped: int = 0


class RevenueSourceItemIn(BaseModel):
    """A single sub row (revenue item) supplied when creating/editing a
    Revenue Source. Any sub rows with a blank description are ignored by
    the endpoints below, so the frontend can freely send/keep placeholder
    rows without them being persisted."""
    id: Optional[int] = None  # present on items coming from an existing source; ignored on write, kept for round-tripping
    description: str
    amount: Union[float, int, str] = 0


class RevenueSourceItemOut(BaseModel):
    id: int
    description: str
    amount: float

    class Config:
        from_attributes = True


class RevenueSourceIn(BaseModel):
    work_plan_id: int
    pbs_fund_code: Optional[str] = None
    source_of_financing_name: str
    functional_definition: Optional[str] = None
    # Used only when `items` is empty/omitted — see RevenueSource model docstring.
    approved_budget_amount: Union[float, int, str] = 0
    items: Optional[List[RevenueSourceItemIn]] = None


class RevenueSourceUpdate(BaseModel):
    pbs_fund_code: Optional[str] = None
    source_of_financing_name: Optional[str] = None
    functional_definition: Optional[str] = None
    approved_budget_amount: Optional[Union[float, int, str]] = None
    # When provided (even as an empty list), this REPLACES all existing sub
    # rows for the revenue source. Omit the field entirely to leave the
    # current sub rows untouched.
    items: Optional[List[RevenueSourceItemIn]] = None


class RevenueSourceOut(BaseModel):
    id: int
    work_plan_id: int
    pbs_fund_code: Optional[str] = None
    source_of_financing_name: str
    functional_definition: Optional[str] = None
    # Auto-derived: equals category_total whenever sub rows exist, otherwise
    # falls back to the manually-entered amount. Kept alongside
    # category_total (identical value) so the Summary of Sources of Revenue
    # table on the frontend — which reads approved_budget_amount — updates
    # automatically without any frontend changes.
    approved_budget_amount: float
    category_total: float
    items: List[RevenueSourceItemOut] = []

    class Config:
        from_attributes = True


class RevenueSourceImportResult(BaseModel):
    created: int
    skipped: int
    errors: List[str] = []


class RevenueSourceClearResult(BaseModel):
    deleted: int


class ActivityIn(BaseModel):
    budget_code_id: int
    name: str
    quarter: str = "Q1"


class ActivityOut(ActivityIn):
    id: int
    is_active: bool
    class Config:
        from_attributes = True


class RequisitionLineItemIn(BaseModel):
    item_no: int
    description: str
    units: Optional[str] = None
    qty: Optional[float] = None
    rate: Optional[float] = None
    amount: float = 0


class VoucherDataIn(BaseModel):
    voucher_no: Optional[str] = None
    budget_output_code: Optional[str] = None
    pv_reference_no: Optional[str] = None
    dr_to: Optional[str] = None
    cheque_no: Optional[str] = None
    address: Optional[str] = None
    voucher_date: Optional[str] = None
    ledger_folio: Optional[str] = None
    charge_date: Optional[str] = None
    authority: Optional[str] = None
    approved_vote: Optional[str] = None
    account_no: Optional[str] = None
    approved_estimate: Optional[str] = None
    cheque_instruction_no: Optional[str] = None
    payment_day: Optional[str] = None
    payment_month_year: Optional[str] = None
    entered_vote_book_date: Optional[str] = None
    verified_by_date: Optional[str] = None
    passed_payment_date: Optional[str] = None
    inter_dept_clearance: Optional[str] = None
    program_of_estimate: Optional[str] = None
    sub_program: Optional[str] = None
    item: Optional[str] = None


class RequisitionIn(BaseModel):
    department_id: Optional[int] = None
    budget_code_id: Optional[int] = None
    activity_id: Optional[int] = None
    subject: Optional[str] = None
    financial_year: Optional[str] = None
    quarter: Optional[str] = None
    requester_name: Optional[str] = None
    requester_position: Optional[str] = None
    requester_mobile: Optional[str] = None
    budget_output_code_text: Optional[str] = None
    payment_voucher_number: Optional[str] = None
    line_items: List[RequisitionLineItemIn]
    voucher: Optional[VoucherDataIn] = None


class ApprovalActionIn(BaseModel):
    action: str  # approve / reject / return
    comments: Optional[str] = None


class AccountabilityIn(BaseModel):
    status: str  # verified / flagged / pending
    remarks: Optional[str] = None


# --------------------------------------------------------------------------
# Auth helpers
# --------------------------------------------------------------------------

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = dt.datetime.utcnow() + dt.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    if token is None:
        raise credentials_exception
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.id == int(user_id)).first()
    if user is None or not user.is_active:
        raise credentials_exception
    return user


def require_roles(*roles):
    def checker(user: User = Depends(get_current_user)):
        if user.role not in roles:
            raise HTTPException(status_code=403, detail="You do not have permission to perform this action")
        return user
    return checker


def log_action(db: Session, user_id: Optional[int], action: str, details: str = ""):
    entry = AuditLog(user_id=user_id, action=action, details=details)
    db.add(entry)
    db.commit()


def notify(db: Session, user_id: int, message: str, category: str = "info", requisition_id: Optional[int] = None):
    n = Notification(user_id=user_id, message=message, category=category, link_requisition_id=requisition_id)
    db.add(n)
    db.commit()


def notify_role(db: Session, role: str, message: str, category: str = "info", requisition_id: Optional[int] = None):
    users = db.query(User).filter(User.role == role, User.is_active == True).all()
    for u in users:
        notify(db, u.id, message, category, requisition_id)


# --------------------------------------------------------------------------
# Backblaze B2 storage helpers
# --------------------------------------------------------------------------

async def upload_document_to_b2(file: UploadFile, requisition_id: int) -> str:
    try:
        ext = os.path.splitext(file.filename or "")[1].lower()
        key = f"{B2_DOCUMENTS_FOLDER}/{requisition_id}/{uuid.uuid4().hex}{ext}"
        file_content = await file.read()
        b2_client.put_object(
            Bucket=B2_BUCKET_NAME,
            Key=key,
            Body=file_content,
            ContentType=file.content_type or "application/octet-stream",
        )
        logger.info(f"Document uploaded to B2: {key}")
        return key
    except Exception as e:
        logger.error(f"Error uploading document to B2: {e}")
        raise HTTPException(status_code=500, detail=f"Error uploading document: {str(e)}")


async def upload_signature_to_b2(file: UploadFile, user_id: int) -> str:
    try:
        ext = os.path.splitext(file.filename or "")[1].lower()
        key = f"{B2_SIGNATURES_FOLDER}/{user_id}/{uuid.uuid4().hex}{ext}"
        file_content = await file.read()
        b2_client.put_object(
            Bucket=B2_BUCKET_NAME,
            Key=key,
            Body=file_content,
            ContentType=file.content_type or "image/png",
        )
        logger.info(f"Signature uploaded to B2: {key}")
        return key
    except Exception as e:
        logger.error(f"Error uploading signature to B2: {e}")
        raise HTTPException(status_code=500, detail=f"Error uploading signature: {str(e)}")


def delete_document_from_b2(key: str):
    try:
        b2_client.delete_object(Bucket=B2_BUCKET_NAME, Key=key)
        logger.info(f"Document deleted from B2: {key}")
    except Exception as e:
        logger.warning(f"Error deleting document from B2: {e}")


# --------------------------------------------------------------------------
# App
# --------------------------------------------------------------------------

app = FastAPI(title="KTC-IPFMS API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def seed_data():
    db = SessionLocal()
    try:
        admin_email = os.getenv("ADMIN_EMAIL", "admin@karugutu.town.go.ug")
        admin_password = os.getenv("ADMIN_PASSWORD", "Admin@2026")

        # Departments are never auto-generated by the system — they are only
        # ever created when a user explicitly adds one (Departments page, or
        # the "+ New Department" action). The seeded admin account is created
        # with no department assigned; an administrator can assign one later
        # once real departments have been added.
        existing_admin = db.query(User).filter(User.email == admin_email).first()
        if not existing_admin:
            admin = User(
                full_name="System Administrator",
                email=admin_email,
                hashed_password=hash_password(admin_password),
                role="admin",
                position="System Administrator",
                department_id=None,
            )
            db.add(admin)
            try:
                db.commit()
                log_action(db, None, "system.seed", f"Seeded initial admin account {admin_email}")
            except IntegrityError:
                db.rollback()
                logger.info("Admin user already existed (created concurrently); skipping seed insert.")
    except Exception as exc:  # noqa: BLE001 - never let seeding crash startup
        db.rollback()
        logger.warning("Startup seeding skipped due to error: %s", exc)
    finally:
        db.close()


# ---------------------------- Auth ----------------------------------------

@app.post("/api/auth/login", response_model=Token)
def login(payload: LoginIn, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email).first()
    if not user or not verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="This account has been deactivated")
    # System Administrator accounts are not limited to the "System
    # Administrator" option on the role selector — an admin may sign in
    # picking any role in the dropdown without it being rejected as a
    # mismatch. Every other account must still select its own real role.
    if payload.role and payload.role != user.role and user.role != "admin":
        raise HTTPException(
            status_code=401,
            detail="The role you selected does not match this account. Please choose the correct role and try again."
        )
    token = create_access_token({"sub": str(user.id), "role": user.role})
    log_action(db, user.id, "auth.login", f"{user.email} logged in")
    return Token(access_token=token, role=user.role, full_name=user.full_name, user_id=user.id)


@app.get("/api/auth/me", response_model=UserOut)
def me(user: User = Depends(get_current_user)):
    return user


# ---------------------------- Users ----------------------------------------

@app.get("/api/users", response_model=List[UserOut])
def list_users(db: Session = Depends(get_db), user: User = Depends(require_roles("admin"))):
    return db.query(User).order_by(User.created_at.desc()).all()


@app.post("/api/users", response_model=UserOut)
def create_user(payload: UserCreate, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    if payload.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role supplied")
    if db.query(User).filter(User.email == payload.email).first():
        raise HTTPException(status_code=400, detail="A user with this email already exists")
    new_user = User(
        full_name=payload.full_name,
        email=payload.email,
        hashed_password=hash_password(payload.password),
        role=payload.role,
        department_id=payload.department_id,
        position=payload.position,
        telephone=payload.telephone,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    log_action(db, admin.id, "user.create", f"Created user {new_user.email} ({new_user.role})")
    return new_user


@app.patch("/api/users/{user_id}/toggle-active", response_model=UserOut)
def toggle_user_active(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    target.is_active = not target.is_active
    db.commit()
    db.refresh(target)
    log_action(db, admin.id, "user.toggle_active", f"{target.email} active={target.is_active}")
    return target


@app.patch("/api/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    data = payload.dict(exclude_unset=True)
    if "role" in data and data["role"] and data["role"] not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role supplied")
    if data.get("email") and data["email"] != target.email:
        if db.query(User).filter(User.email == data["email"]).first():
            raise HTTPException(status_code=400, detail="A user with this email already exists")
    password = data.pop("password", None)
    for field, value in data.items():
        setattr(target, field, value)
    if password:
        target.hashed_password = hash_password(password)
    db.commit()
    db.refresh(target)
    log_action(db, admin.id, "user.update", f"Updated user {target.email}")
    return target


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.id == admin.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    if db.query(Requisition).filter(Requisition.requester_id == user_id).count() > 0:
        raise HTTPException(status_code=400, detail="Cannot delete a user who has requisitions on record — disable the account instead")
    db.delete(target)
    db.commit()
    log_action(db, admin.id, "user.delete", f"Deleted user {target.email}")
    return {"ok": True}


# ---------------------------- User Signatures -------------------------------

_ALLOWED_SIGNATURE_EXT = {".png", ".jpg", ".jpeg"}


@app.post("/api/users/me/signature", response_model=UserOut)
async def upload_my_signature(file: UploadFile = File(...), db: Session = Depends(get_db),
                               user: User = Depends(get_current_user)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in _ALLOWED_SIGNATURE_EXT:
        raise HTTPException(status_code=400, detail="Only PNG and JPG images are allowed for a signature")

    old_path = user.signature_path
    new_key = await upload_signature_to_b2(file, user.id)
    user.signature_path = new_key
    db.commit()
    db.refresh(user)

    if old_path:
        delete_document_from_b2(old_path)

    log_action(db, user.id, "user.signature_upload", f"{user.email} uploaded a new signature")
    return user


@app.delete("/api/users/me/signature", response_model=UserOut)
def delete_my_signature(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if user.signature_path:
        old_path = user.signature_path
        user.signature_path = None
        db.commit()
        db.refresh(user)
        delete_document_from_b2(old_path)
        log_action(db, user.id, "user.signature_remove", f"{user.email} removed their signature")
    return user


# ---------------------------- Departments -----------------------------------

@app.get("/api/departments", response_model=List[DepartmentOut])
def list_departments(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return db.query(Department).order_by(Department.name).all()


@app.post("/api/departments", response_model=DepartmentOut)
def create_department(payload: DepartmentIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    if db.query(Department).filter(Department.code == payload.code).first():
        raise HTTPException(status_code=400, detail="Department code already exists")
    if db.query(Department).filter(Department.name == payload.name).first():
        raise HTTPException(status_code=400, detail="Department name already exists")
    dep = Department(**payload.dict())
    db.add(dep)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Department name or code already exists")
    db.refresh(dep)
    log_action(db, admin.id, "department.create", dep.name)
    _invalidate_budget_code_caches()
    return dep


@app.patch("/api/departments/{dep_id}", response_model=DepartmentOut)
def update_department(dep_id: int, payload: DepartmentIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    dep = db.query(Department).filter(Department.id == dep_id).first()
    if not dep:
        raise HTTPException(status_code=404, detail="Department not found")
    if db.query(Department).filter(Department.code == payload.code, Department.id != dep_id).first():
        raise HTTPException(status_code=400, detail="Department code already exists")
    if db.query(Department).filter(Department.name == payload.name, Department.id != dep_id).first():
        raise HTTPException(status_code=400, detail="Department name already exists")
    dep.name = payload.name
    dep.code = payload.code
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Department name or code already exists")
    db.refresh(dep)
    log_action(db, admin.id, "department.update", dep.name)
    _invalidate_budget_code_caches()
    return dep


@app.delete("/api/departments/{dep_id}")
def delete_department(dep_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    dep = db.query(Department).filter(Department.id == dep_id).first()
    if not dep:
        raise HTTPException(status_code=404, detail="Department not found")
    if db.query(User).filter(User.department_id == dep_id).count() > 0:
        raise HTTPException(status_code=400, detail="Cannot delete a department that still has users assigned to it")
    if db.query(BudgetCode).filter(BudgetCode.department_id == dep_id).count() > 0:
        raise HTTPException(status_code=400, detail="Cannot delete a department that still has budget codes assigned to it")
    db.delete(dep)
    db.commit()
    log_action(db, admin.id, "department.delete", dep.name)
    _invalidate_budget_code_caches()
    return {"ok": True}


# ---------------------------- Work Plans ------------------------------------

def _default_table_titles(financial_year: str) -> dict:
    """Fallback headings for the 4 Work Plan & Budget tables, used whenever a
    work plan is created/updated without explicit text for one of them —
    keeps the tables from ever showing a blank heading."""
    fy = (financial_year or "").strip()
    suffix = f" FOR FY {fy}" if fy else ""
    return {
        "title_revenue_summary": f"APPROVED SUMMARY OF THE COUNCIL BUDGET FRAMEWORK PAPER AND PRELIMINARY REVENUE ESTIMATES{suffix}",
        "title_dept_summary": f"APPROVED DEPARTMENTAL SUMMARY OF THE COUNCIL ANNUAL WORK PLAN AND EXPENDITURE ESTIMATES{suffix}",
        "title_revenue_detail": f"APPROVED COUNCIL BUDGET FRAMEWORK PAPER AND PRELIMINARY REVENUE ESTIMATES{suffix}",
        "title_main_table": f"APPROVED COUNCIL ANNUAL WORK PLAN AND EXPENDITURE ESTIMATES{suffix}",
    }


@app.get("/api/workplans", response_model=List[WorkPlanOut])
def list_workplans(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    # Ordered oldest -> newest (by creation order) so the older annual work
    # plan(s) appear at the top of the dropdown and the most recently added
    # work plan appears at the bottom.
    return db.query(WorkPlan).order_by(WorkPlan.id.asc()).all()


@app.post("/api/workplans", response_model=WorkPlanOut)
def create_workplan(payload: WorkPlanIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    data = payload.dict()
    defaults = _default_table_titles(payload.financial_year)
    for field, default_text in defaults.items():
        if not (data.get(field) or "").strip():
            data[field] = default_text
    wp = WorkPlan(**data)
    db.add(wp)
    db.commit()
    db.refresh(wp)
    log_action(db, admin.id, "workplan.create", f"{wp.title} ({wp.financial_year})")
    _invalidate_budget_code_caches()
    return wp


@app.patch("/api/workplans/{wp_id}", response_model=WorkPlanOut)
def update_workplan(wp_id: int, payload: WorkPlanIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    wp = db.query(WorkPlan).filter(WorkPlan.id == wp_id).first()
    if not wp:
        raise HTTPException(status_code=404, detail="Work plan not found")
    defaults = _default_table_titles(payload.financial_year)
    wp.financial_year = payload.financial_year
    wp.title = payload.title
    wp.title_revenue_summary = (payload.title_revenue_summary or "").strip() or defaults["title_revenue_summary"]
    wp.title_dept_summary = (payload.title_dept_summary or "").strip() or defaults["title_dept_summary"]
    wp.title_revenue_detail = (payload.title_revenue_detail or "").strip() or defaults["title_revenue_detail"]
    wp.title_main_table = (payload.title_main_table or "").strip() or defaults["title_main_table"]
    db.commit()
    db.refresh(wp)
    log_action(db, admin.id, "workplan.update", f"{wp.title} ({wp.financial_year})")
    _invalidate_budget_code_caches()
    return wp


@app.delete("/api/workplans/{wp_id}")
def delete_workplan(wp_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    wp = db.query(WorkPlan).filter(WorkPlan.id == wp_id).first()
    if not wp:
        raise HTTPException(status_code=404, detail="Work plan not found")
    if db.query(BudgetCode).filter(BudgetCode.work_plan_id == wp_id).count() > 0:
        raise HTTPException(status_code=400, detail="Cannot delete a work plan that still has budget codes attached to it")
    db.delete(wp)
    db.commit()
    log_action(db, admin.id, "workplan.delete", f"{wp.title} ({wp.financial_year})")
    _invalidate_budget_code_caches()
    return {"ok": True}


def _normalize_revenue_source_fields(pbs_fund_code, source_of_financing_name, functional_definition):
    """A past frontend bug saved the internal category shorthand ('gou',
    'lr', 'mdp') as the Revenue Source name instead of resolving it to the
    full Council-approved wording (e.g. 'Central Government Transfers
    (GoU)'). Any record — old or new — whose name is exactly one of those
    shorthands is displayed with the correct full label instead."""
    key = (source_of_financing_name or "").strip().lower()
    cat = _REVENUE_KEY_TO_FULL.get(key)
    if cat:
        return (
            pbs_fund_code or cat["pbs_fund_code"],
            cat["source_of_financing_name"],
            functional_definition or cat["functional_definition"],
        )
    return pbs_fund_code, source_of_financing_name, functional_definition


def revenue_source_to_out(r: RevenueSource) -> RevenueSourceOut:
    items_out = [
        RevenueSourceItemOut(id=it.id, description=it.description, amount=parse_amount(it.amount))
        for it in (r.items or [])
    ]
    if items_out:
        total = sum(it.amount for it in items_out)
    else:
        total = parse_amount(r.approved_budget_amount)
    pbs_fund_code, source_name, func_def = _normalize_revenue_source_fields(
        r.pbs_fund_code, r.source_of_financing_name, r.functional_definition
    )
    return RevenueSourceOut(
        id=r.id, work_plan_id=r.work_plan_id, pbs_fund_code=pbs_fund_code,
        source_of_financing_name=source_name,
        functional_definition=func_def,
        approved_budget_amount=total,
        category_total=total,
        items=items_out,
    )


def _apply_revenue_items(db: Session, r: RevenueSource, items_in: List[RevenueSourceItemIn]):
    """Replace all sub rows on a revenue source with the given list. Rows
    with a blank description are dropped rather than persisted, so the
    frontend can send placeholder/incomplete rows freely."""
    db.query(RevenueSourceItem).filter(RevenueSourceItem.revenue_source_id == r.id).delete()
    for it in items_in or []:
        desc = (it.description or "").strip()
        if not desc:
            continue
        db.add(RevenueSourceItem(revenue_source_id=r.id, description=desc, amount=parse_amount(it.amount)))


@app.get("/api/revenue-sources", response_model=List[RevenueSourceOut])
def list_revenue_sources(work_plan_id: Optional[int] = None, db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    cache_key = f"revenue_sources:{work_plan_id}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    q = db.query(RevenueSource)
    if work_plan_id:
        q = q.filter(RevenueSource.work_plan_id == work_plan_id)
    # Ordered by id (insertion order) rather than PBS fund code, so rows
    # imported from Excel are displayed in the exact same order they
    # appeared in the workbook, instead of being re-sorted by fund code.
    sources = q.order_by(RevenueSource.id.asc()).all()
    result = [revenue_source_to_out(r) for r in sources]
    _cache_set(cache_key, result)
    return result


@app.post("/api/revenue-sources", response_model=RevenueSourceOut)
def create_revenue_source(payload: RevenueSourceIn, db: Session = Depends(get_db),
                           admin: User = Depends(require_roles("admin"))):
    wp = db.query(WorkPlan).filter(WorkPlan.id == payload.work_plan_id).first()
    if not wp:
        raise HTTPException(status_code=400, detail="Selected work plan does not exist")

    dup_key = _revenue_source_dup_key(payload.pbs_fund_code, payload.source_of_financing_name)
    if dup_key and dup_key in _existing_revenue_source_keys(db, payload.work_plan_id):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot add this entry — a revenue source named '{payload.source_of_financing_name}'"
                   f"{f' under PBS Fund Code {payload.pbs_fund_code}' if payload.pbs_fund_code else ''}"
                   f" already exists in this work plan.",
        )

    norm_fund_code, norm_source_name, norm_func_def = _normalize_revenue_source_fields(
        payload.pbs_fund_code, payload.source_of_financing_name, payload.functional_definition
    )
    r = RevenueSource(
        work_plan_id=payload.work_plan_id,
        pbs_fund_code=norm_fund_code,
        source_of_financing_name=norm_source_name,
        functional_definition=norm_func_def,
        approved_budget_amount=parse_amount(payload.approved_budget_amount),
    )
    db.add(r)
    db.flush()  # get r.id before adding sub rows
    _apply_revenue_items(db, r, payload.items or [])
    db.commit()
    db.refresh(r)
    log_action(db, admin.id, "revenue_source.create", r.source_of_financing_name)
    _invalidate_revenue_source_caches()
    return revenue_source_to_out(r)


@app.delete("/api/revenue-sources/clear", response_model=RevenueSourceClearResult)
def clear_revenue_sources(work_plan_id: int, db: Session = Depends(get_db),
                           admin: User = Depends(require_roles("admin"))):
    """Delete every revenue source (and their sub rows, via cascade) for a
    given work plan in one go — powers the "Clear" button on the Revenue
    Source by Category table so an administrator can wipe the table clean
    before re-entering or re-importing data, instead of removing each
    category one at a time."""
    wp = db.query(WorkPlan).filter(WorkPlan.id == work_plan_id).first()
    if not wp:
        raise HTTPException(status_code=400, detail="Selected work plan does not exist")
    sources = db.query(RevenueSource).filter(RevenueSource.work_plan_id == work_plan_id).all()
    deleted = len(sources)
    for r in sources:
        db.delete(r)  # cascade="all, delete-orphan" on RevenueSource.items removes sub rows too
    db.commit()
    log_action(db, admin.id, "revenue_source.clear_all",
               f"Cleared {deleted} revenue source(s) from work plan #{work_plan_id}")
    _invalidate_revenue_source_caches()
    return RevenueSourceClearResult(deleted=deleted)


@app.patch("/api/revenue-sources/{rev_id}", response_model=RevenueSourceOut)
def update_revenue_source(rev_id: int, payload: RevenueSourceUpdate, db: Session = Depends(get_db),
                           admin: User = Depends(require_roles("admin"))):
    r = db.query(RevenueSource).filter(RevenueSource.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Revenue source not found")
    data = payload.dict(exclude_unset=True)
    items_data = data.pop("items", None)  # None = leave sub rows untouched; [] or [...] = replace them
    if "approved_budget_amount" in data:
        data["approved_budget_amount"] = parse_amount(data["approved_budget_amount"])
    if "source_of_financing_name" in data:
        norm_fund_code, norm_source_name, norm_func_def = _normalize_revenue_source_fields(
            data.get("pbs_fund_code", r.pbs_fund_code),
            data.get("source_of_financing_name"),
            data.get("functional_definition", r.functional_definition),
        )
        data["pbs_fund_code"] = norm_fund_code
        data["source_of_financing_name"] = norm_source_name
        data["functional_definition"] = norm_func_def
    for field, value in data.items():
        setattr(r, field, value)
    if items_data is not None:
        items_in = [RevenueSourceItemIn(**it) for it in items_data]
        _apply_revenue_items(db, r, items_in)
    db.commit()
    db.refresh(r)
    log_action(db, admin.id, "revenue_source.update", r.source_of_financing_name)
    _invalidate_revenue_source_caches()
    return revenue_source_to_out(r)


@app.delete("/api/revenue-sources/{rev_id}")
def delete_revenue_source(rev_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    r = db.query(RevenueSource).filter(RevenueSource.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Revenue source not found")
    name = r.source_of_financing_name
    db.delete(r)  # cascade="all, delete-orphan" on RevenueSource.items removes its sub rows too
    db.commit()
    log_action(db, admin.id, "revenue_source.delete", name)
    _invalidate_revenue_source_caches()
    return {"ok": True}


@app.post("/api/revenue-sources/{rev_id}/items", response_model=RevenueSourceOut)
def add_revenue_source_item(rev_id: int, payload: RevenueSourceItemIn, db: Session = Depends(get_db),
                             admin: User = Depends(require_roles("admin"))):
    """Add a single sub row (revenue item) to an existing revenue source
    category. The Category Total (and therefore the Approved Budget Amount
    shown in the Summary of Sources of Revenue) is recalculated
    automatically, since it is derived from the sub rows on read."""
    r = db.query(RevenueSource).filter(RevenueSource.id == rev_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Revenue source not found")
    desc = (payload.description or "").strip()
    if not desc:
        raise HTTPException(status_code=400, detail="Please provide a description for the revenue item")
    item = RevenueSourceItem(revenue_source_id=r.id, description=desc, amount=parse_amount(payload.amount))
    db.add(item)
    db.commit()
    db.refresh(r)
    log_action(db, admin.id, "revenue_source_item.create", f"{desc} on {r.source_of_financing_name}")
    _invalidate_revenue_source_caches()
    return revenue_source_to_out(r)


@app.delete("/api/revenue-source-items/{item_id}")
def delete_revenue_source_item(item_id: int, db: Session = Depends(get_db),
                                admin: User = Depends(require_roles("admin"))):
    """Remove a single sub row. The parent revenue source's Category Total
    (and the Summary of Sources of Revenue) updates automatically."""
    item = db.query(RevenueSourceItem).filter(RevenueSourceItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Revenue item not found")
    rev_id = item.revenue_source_id
    parent_name = item.revenue_source.source_of_financing_name if item.revenue_source else None
    db.delete(item)
    db.commit()
    log_action(db, admin.id, "revenue_source_item.delete", f"Removed sub row from {parent_name or ('revenue source #' + str(rev_id))}")
    _invalidate_revenue_source_caches()
    return {"ok": True}


_REVENUE_IMPORT_COLUMN_ALIASES = {
    "pbs fund code": "pbs_fund_code",
    "fund code": "pbs_fund_code",
    "source of financing name": "source_of_financing_name",
    "source of financing": "source_of_financing_name",
    "source of funding": "source_of_financing_name",
    "financing source": "source_of_financing_name",
    "functional definition in pbs": "functional_definition",
    "functional definition in pbs (category & item details)": "functional_definition",
    "functional definition": "functional_definition",
    "revenue item": "item_description",
    "revenue item (functional definition)": "item_description",
    "sub row": "item_description",
    "approved estimate": "item_amount",
    "approved estimate (ugx)": "item_amount",
    "approved budget amount": "approved_budget_amount",
    "approved budget amount (ugx)": "approved_budget_amount",
    "approved budget": "approved_budget_amount",
    "subtotal approved budget estimates by revenue source (ugx)": "item_amount",
    "total approved budget estimate by revenue source category (ugx)": "approved_budget_amount",
    "amount": "approved_budget_amount",
    # NEW — matches this workbook's actual column headers (no "Subtotal"/"Total" prefix)
    "approved budget estimates by revenue source (ugx)": "item_amount",
    "approved budget estimate by revenue source category (ugx)": "approved_budget_amount",
}


@app.post("/api/revenue-sources/import", response_model=RevenueSourceImportResult)
async def import_revenue_sources(work_plan_id: int, file: UploadFile = File(...),
                                  db: Session = Depends(get_db),
                                  admin: User = Depends(require_roles("admin"))):
    """Bulk-create Revenue Source rows (and, where present, their sub rows)
    from an uploaded Excel workbook, so revenue sources prepared offline can
    be imported directly instead of being typed in one at a time via the
    Add Revenue Sources form.

    Two supported layouts:
      - A flat sheet with one row per source and an "Approved Budget
        Amount" column (no sub rows) — behaves as before.
      - A "Revenue Entry Table"-style sheet where the same PBS Fund
        Code / Source of Financing Name repeats across consecutive rows,
        each row carrying one "Revenue Item" + "Approved Estimate". These
        are grouped into a single revenue source with one sub row per
        distinct Revenue Item row; the source's own Approved Budget Amount
        column is then ignored in favour of the sub-row total.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel import is not available on this server — the 'openpyxl' package is not installed."
        )

    wp = db.query(WorkPlan).filter(WorkPlan.id == work_plan_id).first()
    if not wp:
        raise HTTPException(status_code=400, detail="Selected work plan does not exist")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx Excel workbook")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}")

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        raise HTTPException(
            status_code=400,
            detail="The uploaded workbook appears to be empty"
        )

    header_row_idx = None
    col_map = {}

    for i, row in enumerate(all_rows[:10]):
        candidate = [_normalize_header_key(h) for h in row]
        candidate_map = {
            idx: _REVENUE_IMPORT_COLUMN_ALIASES[h]
            for idx, h in enumerate(candidate)
            if h in _REVENUE_IMPORT_COLUMN_ALIASES
        }

        if len(candidate_map) >= 2:
            header_row_idx = i
            col_map = candidate_map
            break

    if header_row_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find a header row with recognisable columns (e.g. 'Source of Financing Name') in the first 10 rows of the sheet"
        )

    rows = all_rows[header_row_idx:]

    if "source_of_financing_name" not in col_map.values():
        raise HTTPException(
            status_code=400,
            detail="The workbook must at least include a 'Source of Financing Name' column"
        )

    has_sub_row_columns = "item_description" in col_map.values()

    def _text(v):
        return str(v).strip() if v is not None else None

    created = 0
    skipped = 0
    errors: List[str] = []

    grouped: dict = {}
    order: List[tuple] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        data = {}
        for idx, field in col_map.items():
            data[field] = row[idx] if idx < len(row) else None

        source_name = _text(data.get("source_of_financing_name"))

        if not source_name:
            continuation_text = (
                _text(data.get("item_description"))
                or _text(data.get("functional_definition"))
            )

            if order and continuation_text:
                # Genuine continuation row of the group above it (merged
                # "Source of Financing Name" cell in the source sheet).
                key = order[-1]
            else:
                # FIX: no longer skipped. A row with data but no source
                # name and nothing to tie it to the previous group is
                # still imported — as its own standalone revenue source —
                # so every row in the workbook ends up in the system,
                # instead of quietly disappearing from the import.
                fund_code = _text(data.get("pbs_fund_code"))
                key = (
                    fund_code or "",
                    f"__row{row_idx}__",
                )
                grouped[key] = {
                    "pbs_fund_code": fund_code,
                    "source_of_financing_name": continuation_text or "Unspecified",
                    "functional_definition": _text(data.get("functional_definition")),
                    "approved_budget_amount": data.get("approved_budget_amount"),
                    "items": [],
                }
                order.append(key)
                errors.append(
                    f"Row {row_idx}: Source of Financing Name was blank — imported as 'Unspecified'"
                    if not continuation_text else
                    f"Row {row_idx}: Source of Financing Name was blank — imported as its own row"
                )

        else:
            if _normalize_dept_name(source_name) in _SUBTOTAL_MARKERS:
                continue

            fund_code = _text(data.get("pbs_fund_code"))
            key = (fund_code or "", _normalize_dept_name(source_name))

            if key not in grouped:
                grouped[key] = {
                    "pbs_fund_code": fund_code,
                    "source_of_financing_name": source_name,
                    "functional_definition": _text(data.get("functional_definition")),
                    "approved_budget_amount": data.get("approved_budget_amount"),
                    "items": [],
                }
                order.append(key)

        entry = grouped[key]

        item_desc = _text(data.get("item_description"))
        if not item_desc and not source_name:
            item_desc = _text(data.get("functional_definition"))

        if item_desc:
            amt, ok, original = parse_amount_verbose(data.get("item_amount"))

            if not ok:
                errors.append(
                    f"Row {row_idx}: could not read '{original}' as a number for Approved Estimate — treated as 0"
                )

            entry["items"].append(
                {
                    "description": item_desc,
                    "amount": amt,
                }
            )

    # Tracks every (PBS Fund Code, Source of Financing Name) pair already on
    # file for this work plan (seeded from the database, then grown as rows
    # are imported) so a source already present — or repeated twice within
    # the same workbook — is rejected instead of creating a duplicate entry.
    existing_revenue_source_keys = _existing_revenue_source_keys(db, work_plan_id)

    for key in order:
        entry = grouped[key]

        # Reject this group outright if it's already in the database (or
        # was already imported earlier in this same workbook) — same rule
        # the manual "Add Revenue Source" form enforces, applied here to
        # every imported row/group.
        dup_key = _revenue_source_dup_key(entry["pbs_fund_code"], entry["source_of_financing_name"])
        if dup_key and dup_key in existing_revenue_source_keys:
            skipped += 1
            errors.append(
                f"'{entry['source_of_financing_name']}' skipped — already exists in this work plan"
            )
            continue

        amount, ok, original = parse_amount_verbose(
            entry.get("approved_budget_amount")
        )

        if not ok:
            errors.append(
                f"Could not read '{original}' as a number for Approved Budget Amount on '{entry['source_of_financing_name']}' — treated as 0"
            )

        r = RevenueSource(
            work_plan_id=work_plan_id,
            pbs_fund_code=entry["pbs_fund_code"],
            source_of_financing_name=entry["source_of_financing_name"],
            functional_definition=entry["functional_definition"],
            approved_budget_amount=amount,
        )

        db.add(r)
        db.flush()

        for it in entry["items"]:
            db.add(
                RevenueSourceItem(
                    revenue_source_id=r.id,
                    description=it["description"],
                    amount=it["amount"],
                )
            )

        created += 1
        if dup_key:
            existing_revenue_source_keys.add(dup_key)

    db.commit()

    log_action(
        db,
        admin.id,
        "revenue_source.import",
        f"Imported {created} revenue source(s) into work plan #{work_plan_id} from {file.filename} ({skipped} skipped)",
    )

    _invalidate_revenue_source_caches()

    return RevenueSourceImportResult(
        created=created,
        skipped=skipped,
        errors=errors[:30],
    )

# ---------------------------- Budget Codes ----------------------------------

def _dept_label(dept: Optional["Department"]) -> Optional[str]:
    """Renders a department using just its name. Strict separation of
    concern: the PBS department code is a distinct identifier from the
    department name and is intentionally omitted everywhere a department is
    shown to the user (Departments page's own Department column, user
    records, requisitions, dept-output charts, etc).

    The exceptions are the Annual Work Plan & Budget Estimates table's
    Department column and the Department Budget Summary table (both on
    screen and in the PDF report), plus every dropdown that lists
    departments (Work Plan filter, Budget Code form, Requisition form, User
    form) — those all show "<code>: <name>" (e.g. "090: Community Based
    Services") via _dept_code_and_name() below.
    """
    if not dept:
        return None
    return dept.name


def _dept_code_and_name(dept: Optional["Department"]) -> Optional[str]:
    """Renders "<code>: <name>" (e.g. "090: Community Based Services") for
    the Annual Work Plan & Budget Estimates table's Department column, the
    Department Budget Summary table, and department dropdowns. Do not use
    this for the Departments page's own Department column — that shows
    department name alone via _dept_label()."""
    if not dept:
        return None
    if not dept.code:
        return dept.name
    return f"{dept.code}: {dept.name}"


def budget_code_to_out(bc: BudgetCode, committed_override: Optional[float] = None) -> BudgetCodeOut:
    allocated = bc.allocated_amount
    committed = committed_override if committed_override is not None else bc.committed_amount
    return BudgetCodeOut(
        id=bc.id, work_plan_id=bc.work_plan_id, department_id=bc.department_id,
        department_name=_dept_label(bc.department),
        department_code_and_name=_dept_code_and_name(bc.department),
        service_area=bc.service_area,
        code=bc.code, output_description=bc.output_description, programme=bc.programme,
        sub_programme=bc.sub_programme,
        piap_output_description=bc.piap_output_description,
        piap_output_indicator=bc.piap_output_indicator,
        unit_of_measure=bc.unit_of_measure,
        baseline_value=parse_amount(bc.baseline_value), baseline_note=bc.baseline_note,
        planned_target=parse_amount(bc.planned_target), target_note=bc.target_note,
        actual_output=bc.actual_output,
        q1_amount=parse_amount(bc.q1_amount), q2_amount=parse_amount(bc.q2_amount),
        q3_amount=parse_amount(bc.q3_amount), q4_amount=parse_amount(bc.q4_amount),
        funding_source=bc.funding_source,
        responsible_party=bc.responsible_party,
        allocated_amount=allocated, committed_amount=committed,
        available_balance=allocated - committed,
    )


def _bulk_committed_amounts(db: Session, budget_code_ids: List[int]) -> dict:
    """Committed amount (sum of non-draft/rejected/returned requisitions)
    for a whole batch of budget codes in a single grouped query, instead of
    one query per budget code. This is the main fix for the Work Plan &
    Budget table being slow to load with many rows."""
    if not budget_code_ids:
        return {}
    rows = (
        db.query(Requisition.budget_code_id, func.sum(Requisition.amount_requested))
        .filter(
            Requisition.budget_code_id.in_(budget_code_ids),
            Requisition.status.notin_(["rejected", "returned", "draft"]),
        )
        .group_by(Requisition.budget_code_id)
        .all()
    )
    return {bc_id: (amt or 0.0) for bc_id, amt in rows}


@app.get("/api/budget-codes", response_model=List[BudgetCodeOut])
def list_budget_codes(work_plan_id: Optional[int] = None, department_id: Optional[int] = None,
                       search: Optional[str] = None,
                       db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    cache_key = f"budget_codes:{work_plan_id}:{department_id}:{(search or '').strip().lower()}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    # joinedload(BudgetCode.department) is the other half of the fix (along
    # with _bulk_committed_amounts above it): budget_code_to_out() reads
    # bc.department for every row (department_name / department_code_and_name).
    # Without eager-loading, that's a lazy-loaded relationship, so every row
    # triggered its own extra department SELECT — an N+1 pattern just like the
    # old committed_amount one, and the real reason this endpoint could still
    # crawl (or effectively hang, since the frontend fetch has no timeout) on
    # a work plan with a few hundred imported rows even after that first fix.
    # One joinedload turns "N extra queries" into "0 extra queries" by
    # fetching departments in the same query via a SQL JOIN.
    q = db.query(BudgetCode).options(joinedload(BudgetCode.department))
    if work_plan_id:
        q = q.filter(BudgetCode.work_plan_id == work_plan_id)
    if department_id:
        q = q.filter(BudgetCode.department_id == department_id)
    if search:
        like = f"%{search}%"
        q = q.filter(BudgetCode.output_description.ilike(like) | BudgetCode.code.ilike(like))
    # Ordered by id (insertion order) rather than code, so rows imported
    # from Excel are displayed in the exact same order they appeared in the
    # workbook, instead of being re-sorted alphabetically by code.
    codes = q.order_by(BudgetCode.id.asc()).all()

    committed_map = _bulk_committed_amounts(db, [bc.id for bc in codes])
    result = [budget_code_to_out(bc, committed_map.get(bc.id, 0.0)) for bc in codes]

    _cache_set(cache_key, result)
    return result


_BUDGET_CODE_NUMERIC_FIELDS = ("baseline_value", "planned_target", "q1_amount", "q2_amount", "q3_amount", "q4_amount")


@app.post("/api/budget-codes", response_model=BudgetCodeOut)
def create_budget_code(payload: BudgetCodeIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    data = payload.dict()
    for f in _BUDGET_CODE_NUMERIC_FIELDS:
        data[f] = parse_amount(data[f])

    dup_key = _budget_code_dup_key(data.get("code"), data.get("output_description"))
    if dup_key and dup_key in _existing_budget_code_keys(db, data["work_plan_id"]):
        raise HTTPException(status_code=400, detail=f"Cannot add this entry — {_budget_code_dup_message(dup_key)}.")

    bc = BudgetCode(**data)
    db.add(bc)
    db.commit()
    db.refresh(bc)
    log_action(db, admin.id, "budget_code.create", f"{bc.code} - {bc.output_description}")
    _invalidate_budget_code_caches()
    return budget_code_to_out(bc)


@app.delete("/api/budget-codes/clear", response_model=BudgetCodeClearResult)
def clear_budget_codes(work_plan_id: int, db: Session = Depends(get_db),
                        admin: User = Depends(require_roles("admin"))):
    """Delete every Activity & Budget Estimate row for a given work plan in
    one go — powers the "Clear" button on the Annual Work Plan table so an
    administrator can wipe the table clean before re-entering or
    re-importing data. Rows that already have requisitions raised against
    them are left in place (same protection as the single-row delete
    endpoint) and reported back as skipped rather than blocking the whole
    operation."""
    wp = db.query(WorkPlan).filter(WorkPlan.id == work_plan_id).first()
    if not wp:
        raise HTTPException(status_code=400, detail="Selected work plan does not exist")

    codes = db.query(BudgetCode).filter(BudgetCode.work_plan_id == work_plan_id).all()
    deleted = 0
    skipped = 0
    for bc in codes:
        if db.query(Requisition).filter(Requisition.budget_code_id == bc.id).count() > 0:
            skipped += 1
            continue
        db.query(Activity).filter(Activity.budget_code_id == bc.id).delete()
        db.delete(bc)
        deleted += 1

    db.commit()
    log_action(db, admin.id, "budget_code.clear_all",
               f"Cleared {deleted} budget estimate row(s) from work plan #{work_plan_id} ({skipped} skipped — have requisitions on record)")
    _invalidate_budget_code_caches()
    return BudgetCodeClearResult(deleted=deleted, skipped=skipped)


@app.patch("/api/budget-codes/{bc_id}", response_model=BudgetCodeOut)
def update_budget_code(bc_id: int, payload: BudgetCodeUpdate, db: Session = Depends(get_db),
                        admin: User = Depends(require_roles("admin"))):
    bc = db.query(BudgetCode).filter(BudgetCode.id == bc_id).first()
    if not bc:
        raise HTTPException(status_code=404, detail="Budget code not found")
    data = payload.dict(exclude_unset=True)
    for f in _BUDGET_CODE_NUMERIC_FIELDS:
        if f in data:
            data[f] = parse_amount(data[f])
    for field, value in data.items():
        setattr(bc, field, value)
    db.commit()
    db.refresh(bc)
    log_action(db, admin.id, "budget_code.update", f"{bc.code} - {bc.output_description}")
    _invalidate_budget_code_caches()
    return budget_code_to_out(bc)


@app.delete("/api/budget-codes/{bc_id}")
def delete_budget_code(bc_id: int, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    bc = db.query(BudgetCode).filter(BudgetCode.id == bc_id).first()
    if not bc:
        raise HTTPException(status_code=404, detail="Budget code not found")
    if db.query(Requisition).filter(Requisition.budget_code_id == bc_id).count() > 0:
        raise HTTPException(status_code=400, detail="Cannot delete a budget code that has requisitions raised against it")
    db.query(Activity).filter(Activity.budget_code_id == bc_id).delete()
    db.delete(bc)
    db.commit()
    log_action(db, admin.id, "budget_code.delete", f"{bc.code} - {bc.output_description}")
    _invalidate_budget_code_caches()
    return {"ok": True}


# ---- Excel import ----------------------------------------------------------
#
# FIXES APPLIED IN THIS VERSION (see inline comments below for details):
#
#   1. Header matching now collapses ALL whitespace (including embedded
#      newlines, e.g. "Q1 \n(UGX)" as produced by Excel's alt-enter line
#      wraps in a header cell) before comparing against the alias table.
#      Previously "Q1 \n(UGX)".strip().lower() == "q1 \n(ugx)", which never
#      matched the "q1 (ugx)" alias key, so Q1-Q4 (and therefore the derived
#      Total Budget) were silently mapped to nothing and always stored as 0
#      — regardless of what was actually typed in the workbook.
#
#   2. Department matching is now tolerant of "&" vs "and", extra spacing,
#      and case differences (the same council's own workbook mixes
#      "Administration and Support Services" and "Administration & Support
#      Services" across rows). A department that still can't be matched
#      after normalising is auto-created (import is admin-only, so this is
#      safe) instead of the row being silently skipped — this is what was
#      causing "only one department imported, everything after is left
#      out": once the normalized name stopped matching, every remaining row
#      in the workbook (including new departments and quarter figures) was
#      rejected as "department not found".
#
#   3. Rows that are not real data — repeated header rows (Excel workbooks
#      exported per-page often repeat the header every ~15-20 rows) and
#      "Sub Total"/"Total" summary rows — are now recognised and skipped
#      quietly (not counted as warnings/errors), instead of being reported
#      as confusing "department not found" errors.

_IMPORT_COLUMN_ALIASES = {
    "department": "department",
    "service area": "service_area",
    "programme": "programme",
    "program": "programme",
    "sub programme": "sub_programme",
    "sub-programme": "sub_programme",
    "sub program": "sub_programme",
    "budget output code": "code",
    "budget output": "output_description",
    "budget output description": "output_description",
    "piap output description": "piap_output_description",
    "piap output indicator": "piap_output_indicator",
    "unit of measure": "unit_of_measure",
    "baseline value": "baseline_value",
    "baseline": "baseline_value",
    "planned target": "planned_target",
    "target": "planned_target",
    "actual output": "actual_output",
    "q1(ugx)": "q1_amount", "q1 (ugx)": "q1_amount", "q1": "q1_amount",
    "q2(ugx)": "q2_amount", "q2 (ugx)": "q2_amount", "q2": "q2_amount",
    "q3(ugx)": "q3_amount", "q3 (ugx)": "q3_amount", "q3": "q3_amount",
    "q4(ugx)": "q4_amount", "q4 (ugx)": "q4_amount", "q4": "q4_amount",
    "total budget (ugx)": "_total_budget_ignored",
    "total budget": "_total_budget_ignored",
    "funding source": "funding_source",
    "revenue source": "funding_source",
    "responsible party": "responsible_party",
}

_NUMERIC_FIELD_LABELS = {
    "baseline_value": "Baseline Value",
    "planned_target": "Planned Target",
    "q1_amount": "Q1 (UGX)",
    "q2_amount": "Q2 (UGX)",
    "q3_amount": "Q3 (UGX)",
    "q4_amount": "Q4 (UGX)",
}

# Collapses ANY run of whitespace — spaces, tabs, and (crucially) the
# embedded newlines Excel inserts when a header cell uses alt-enter line
# wraps, e.g. "Q1 \n(UGX)" — down to a single space, so header text always
# normalises to the same key regardless of how it was line-wrapped in the
# source spreadsheet.
_WHITESPACE_RE = re.compile(r"\s+")


def _normalize_header_key(h) -> str:
    if h is None:
        return ""
    return _WHITESPACE_RE.sub(" ", str(h)).strip().lower()


# Normalises a department name for matching: case-insensitive, "&" treated
# the same as "and", and all whitespace collapsed. This is what lets
# "Administration & Support Services" match an existing department already
# stored as "Administration and Support Services". Also reused (loosely)
# by the revenue-source importer to detect subtotal/total marker rows.
def _normalize_dept_name(name: Optional[str]) -> str:
    if not name:
        return ""
    n = _WHITESPACE_RE.sub(" ", str(name)).strip().lower()
    n = re.sub(r"\s*&\s*", " and ", n)
    n = _WHITESPACE_RE.sub(" ", n).strip()
    return n


# Rows whose "Department" cell is one of these (after normalising) are not
# real budget entries — they're page-footer subtotal/total rows carried
# over from the source workbook's print layout — and should be skipped
# quietly rather than reported as import errors.
_SUBTOTAL_MARKERS = {"sub total", "subtotal", "total", "grand total"}


# --------------------------------------------------------------------------
# Duplicate-entry guards
# --------------------------------------------------------------------------
# Stops the same record from ending up in the database twice — whether it
# arrives via the Excel importer or is typed in manually through the "Add"
# forms. Both paths funnel through the same key/lookup helpers below, so a
# row that's already on file is rejected the same way either way.
#
#   - Budget Codes: two rows are considered the same entry if they share a
#     Budget Output Code (case/space insensitive) within the same work
#     plan. When the code is blank, the Budget Output Description is used
#     as the fallback identity instead. If both are blank there's nothing
#     reliable to key on, so that one row is left alone (matches the
#     existing "imported with a blank code/description" warning path).
#   - Revenue Sources: two rows are considered the same entry if they share
#     a (PBS Fund Code, Source of Financing Name) pair (case/space
#     insensitive) within the same work plan.

def _budget_code_dup_key(code: Optional[str], output_description: Optional[str]) -> Optional[tuple]:
    norm_code = _normalize_header_key(code)
    if norm_code:
        return ("code", norm_code)
    norm_desc = _normalize_header_key(output_description)
    if norm_desc:
        return ("desc", norm_desc)
    return None


def _existing_budget_code_keys(db: Session, work_plan_id: int) -> set:
    rows = (
        db.query(BudgetCode.code, BudgetCode.output_description)
        .filter(BudgetCode.work_plan_id == work_plan_id)
        .all()
    )
    keys = set()
    for code, desc in rows:
        key = _budget_code_dup_key(code, desc)
        if key:
            keys.add(key)
    return keys


def _budget_code_dup_message(dup_key: tuple) -> str:
    field = "Budget Output Code" if dup_key[0] == "code" else "Budget Output Description"
    return f"a budget estimate row with this {field} already exists in this work plan"


def _revenue_source_dup_key(pbs_fund_code: Optional[str], source_of_financing_name: Optional[str]) -> Optional[tuple]:
    norm_name = _normalize_dept_name(source_of_financing_name)
    if not norm_name:
        return None
    return (_normalize_header_key(pbs_fund_code), norm_name)


def _existing_revenue_source_keys(db: Session, work_plan_id: int) -> set:
    rows = (
        db.query(RevenueSource.pbs_fund_code, RevenueSource.source_of_financing_name)
        .filter(RevenueSource.work_plan_id == work_plan_id)
        .all()
    )
    keys = set()
    for fund_code, name in rows:
        key = _revenue_source_dup_key(fund_code, name)
        if key:
            keys.add(key)
    return keys


def _generate_department_code(name: str, existing_codes: set) -> str:
    """Best-effort short code derived from a department's name, for when a
    brand-new department has to be auto-created during import (the source
    workbook only carries the name, not a short code). Falls back to a
    numbered suffix if the derived code collides with one already in use."""
    words = re.findall(r"[A-Za-z0-9]+", name or "")
    if not words:
        base = "DEPT"
    else:
        base = "".join(w[0] for w in words[:6]).upper()
        if len(base) < 2:
            base = (words[0][:4]).upper()
    base = base[:12] or "DEPT"
    code = base
    n = 1
    while code in existing_codes:
        n += 1
        code = f"{base}{n}"
    return code


@app.post("/api/budget-codes/import", response_model=BudgetCodeImportResult)
async def import_budget_codes(work_plan_id: int, file: UploadFile = File(...),
                               db: Session = Depends(get_db),
                               admin: User = Depends(require_roles("admin"))):
    """
    Bulk-create Budget Estimates rows from an uploaded Excel workbook so a
    user can prepare the "New Budget Estimates Data Entry Form" data offline
    (e.g. in the Council's existing spreadsheet template) and have every row
    populate directly into the system, instead of typing each one in
    manually.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel import is not available on this server — the 'openpyxl' package is not installed."
        )

    wp = db.query(WorkPlan).filter(WorkPlan.id == work_plan_id).first()
    if not wp:
        raise HTTPException(status_code=400, detail="Selected work plan does not exist")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx Excel workbook")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded workbook appears to be empty")

    # FIX (1): normalize header text (collapsing embedded newlines like the
    # "Q1 \n(UGX)" case) before looking it up in the alias table.
    header = [_normalize_header_key(h) for h in rows[0]]
    col_map = {}  # column index -> field name
    for idx, h in enumerate(header):
        field = _IMPORT_COLUMN_ALIASES.get(h)
        if field:
            col_map[idx] = field
    dept_col_idx = next((idx for idx, f in col_map.items() if f == "department"), None)
    code_col_idx = next((idx for idx, f in col_map.items() if f == "code"), None)

    if "output_description" not in col_map.values() or "code" not in col_map.values():
        raise HTTPException(
            status_code=400,
            detail="The workbook must at least include 'Budget Output Code' and 'Budget Output Description' columns"
        )

    # FIX (2): build the department lookup keyed by a normalized name (case
    # / "&" vs "and" / whitespace insensitive) so workbook rows using either
    # spelling resolve to the same existing department.
    all_departments = db.query(Department).all()
    departments_by_normalized_name = {_normalize_dept_name(d.name): d for d in all_departments}
    departments_by_code = {_normalize_header_key(d.code): d for d in all_departments}
    existing_dept_codes = {d.code for d in all_departments}
    departments_created = 0

    created = 0
    skipped = 0
    errors: List[str] = []

    # Tracks every Budget Output Code / Budget Output Description already
    # on file for this work plan (seeded from the database, then grown as
    # rows are imported) so a row already present — or repeated twice
    # within the same workbook — is rejected instead of creating a
    # duplicate entry.
    existing_budget_code_keys = _existing_budget_code_keys(db, work_plan_id)

    def _num(v, field_key: str = None, row_idx: int = None, row_warnings: list = None):
        value, ok, original = parse_amount_verbose(v)
        if not ok and row_warnings is not None:
            label = _NUMERIC_FIELD_LABELS.get(field_key, field_key or "value")
            loc = f"Row {row_idx}: " if row_idx else ""
            row_warnings.append(f"{loc}could not read '{original}' as a number for {label} — treated as 0")
        return value

    # Baseline Value / Planned Target are frequently narrative or
    # multi-part text in real PIAP work plans (e.g. "25% of Council area
    # mapped for vectors", or a per-quarter list like "12; 4; 4; 4; 4; 4;
    # 4"), not a single clean number. Rather than zeroing these out and
    # losing the detail, we pull out a best-effort leading number (for
    # existing numeric aggregates/charts) and keep the full original text
    # in a companion *_note field. This does NOT raise a row warning —
    # unlike Q1-Q4 amounts, non-numeric baseline/target text is expected
    # and normal, not something to flag as a problem.
    def _num_with_note(v):
        value, was_text = parse_leading_number(v)
        note = None
        if was_text:
            _, _, original = parse_amount_verbose(v)
            note = original or None
        return value, note

    def _text(v):
        return str(v).strip() if v is not None else None

    # FIX (4): every non-blank row in the workbook is now imported — rows
    # are no longer skipped just because the Department, Budget Output
    # Code, or Budget Output Description cell is empty. Excel workbooks
    # routinely leave those cells blank on continuation rows (merged
    # cells that only carry a value on the first row of a group), so
    # requiring them on every single row silently dropped legitimate data.
    # Department (and, as a light touch, the other grouping columns) now
    # simply carries forward from the last row that had a value, matching
    # how the source workbook visually groups rows under one heading.
    last_dept_name_raw: Optional[str] = None
    last_service_area: Optional[str] = None
    last_programme: Optional[str] = None
    last_sub_programme: Optional[str] = None

    for row_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # truly blank row — nothing in it to import

        data = {}
        for idx, field in col_map.items():
            data[field] = row[idx] if idx < len(row) else None

        dept_name_raw = _text(data.get("department"))

        # FIX (3a): silently skip repeated header rows. These occur when the
        # source workbook was formatted for printing and the header row was
        # repeated every page; the whole row (department/code/etc.) matches
        # the actual header text, so it is not a data row at all.
        raw_dept_cell = row[dept_col_idx] if dept_col_idx is not None and dept_col_idx < len(row) else None
        raw_code_cell = row[code_col_idx] if code_col_idx is not None and code_col_idx < len(row) else None
        if _normalize_header_key(raw_dept_cell) == "department" or _normalize_header_key(raw_code_cell) == "budget output code":
            continue

        # FIX (3b): silently skip "Sub Total" / "Total" summary rows carried
        # over from the workbook's print layout — these aren't real budget
        # output entries and have no code/description of their own.
        if _normalize_dept_name(dept_name_raw) in _SUBTOTAL_MARKERS:
            continue

        row_warnings: List[str] = []

        code = _text(data.get("code"))
        output_description = _text(data.get("output_description"))
        if not code:
            row_warnings.append(f"Row {row_idx}: Budget Output Code was blank — imported with a blank code")
        if not output_description:
            row_warnings.append(f"Row {row_idx}: Budget Output Description was blank — imported with a blank description")

        # Reject this row outright if it's already in the database (or was
        # already imported earlier in this same workbook) — same rule the
        # manual "Add" form enforces, applied here to every imported row.
        dup_key = _budget_code_dup_key(code, output_description)
        if dup_key and dup_key in existing_budget_code_keys:
            skipped += 1
            errors.append(f"Row {row_idx}: skipped — {_budget_code_dup_message(dup_key)}")
            continue

        service_area = _text(data.get("service_area")) or last_service_area
        programme = _text(data.get("programme")) or last_programme
        sub_programme = _text(data.get("sub_programme")) or last_sub_programme
        last_service_area = service_area or last_service_area
        last_programme = programme or last_programme
        last_sub_programme = sub_programme or last_sub_programme

        if not dept_name_raw:
            # Continuation row under the same department heading as the
            # last row that specified one (a merged-cell group in the
            # source sheet). Departments are never auto-generated by the
            # system, so if no department has been seen yet in this sheet
            # the row is skipped and flagged rather than filed under a
            # placeholder department.
            dept_name_raw = last_dept_name_raw
            if not dept_name_raw:
                skipped += 1
                errors.append(f"Row {row_idx}: no department specified anywhere above this row — skipped. Add the department first, then re-import.")
                continue
        last_dept_name_raw = dept_name_raw

        normalized_dept = _normalize_dept_name(dept_name_raw)
        dept = departments_by_normalized_name.get(normalized_dept)
        if not dept:
            # The PBS workbook often writes the department cell as
            # "<code>: <name>" (e.g. "090: Community Based Services"),
            # while a Department record stores the code and name in
            # separate fields. Strip a leading "<code>:" / "<code> -" /
            # "<code>." style prefix and retry the name match, and also
            # try matching the prefix directly against a department's
            # code, before concluding the department truly doesn't exist.
            code_prefix_match = re.match(r"^\s*([A-Za-z0-9]+)\s*[:.\-]\s*(.+)$", dept_name_raw)
            if code_prefix_match:
                prefix, remainder = code_prefix_match.group(1), code_prefix_match.group(2)
                dept = departments_by_normalized_name.get(_normalize_dept_name(remainder))
                if not dept:
                    dept = departments_by_code.get(_normalize_header_key(prefix))
        if not dept:
            # Departments are only ever created when a user explicitly adds
            # one (Departments page). The importer no longer creates them on
            # the fly — a row referencing an unknown department is skipped
            # and flagged so it can be re-imported once that department
            # has been added.
            skipped += 1
            errors.append(f"Row {row_idx}: department '{dept_name_raw}' does not exist — skipped. Add it under Departments first, then re-import.")
            continue

        baseline_value, baseline_note = _num_with_note(data.get("baseline_value"))
        planned_target, target_note = _num_with_note(data.get("planned_target"))

        bc = BudgetCode(
            work_plan_id=work_plan_id,
            department_id=dept.id,
            service_area=service_area,
            code=code or "",
            output_description=output_description or "",
            programme=programme,
            sub_programme=sub_programme,
            piap_output_description=_text(data.get("piap_output_description")),
            piap_output_indicator=_text(data.get("piap_output_indicator")),
            unit_of_measure=_text(data.get("unit_of_measure")),
            baseline_value=baseline_value,
            baseline_note=baseline_note,
            planned_target=planned_target,
            target_note=target_note,
            actual_output=_text(data.get("actual_output")),
            q1_amount=_num(data.get("q1_amount"), "q1_amount", row_idx, row_warnings),
            q2_amount=_num(data.get("q2_amount"), "q2_amount", row_idx, row_warnings),
            q3_amount=_num(data.get("q3_amount"), "q3_amount", row_idx, row_warnings),
            q4_amount=_num(data.get("q4_amount"), "q4_amount", row_idx, row_warnings),
            funding_source=_text(data.get("funding_source")) or "Local Revenue",
            responsible_party=_text(data.get("responsible_party")),
        )
        db.add(bc)
        created += 1
        if dup_key:
            existing_budget_code_keys.add(dup_key)
        errors.extend(row_warnings)

    db.commit()
    log_action(db, admin.id, "budget_code.import",
               f"Imported {created} row(s) into work plan #{work_plan_id} from {file.filename} "
               f"({skipped} skipped, {departments_created} department(s) auto-created)")
    _invalidate_budget_code_caches()
    return BudgetCodeImportResult(created=created, skipped=skipped, departments_created=departments_created,
                                   errors=errors[:30], total_warnings=len(errors))


# ---------------------------- Activities ------------------------------------

@app.get("/api/activities", response_model=List[ActivityOut])
def list_activities(budget_code_id: Optional[int] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    q = db.query(Activity).filter(Activity.is_active == True)
    if budget_code_id:
        q = q.filter(Activity.budget_code_id == budget_code_id)
    return q.all()


@app.post("/api/activities", response_model=ActivityOut)
def create_activity(payload: ActivityIn, db: Session = Depends(get_db), admin: User = Depends(require_roles("admin"))):
    act = Activity(**payload.dict())
    db.add(act)
    db.commit()
    db.refresh(act)
    log_action(db, admin.id, "activity.create", act.name)
    return act


# ---------------------------- Requisitions ----------------------------------

def _dept_ref_slug(department_name: Optional[str]) -> str:
    """Turns a department name into the short upper-case token used in the
    reference number, e.g. "Finance & Administration" -> "FINANCE-ADMINISTRATION"."""
    name = (department_name or "GENERAL").strip().upper()
    slug = re.sub(r"[^A-Z0-9]+", "-", name).strip("-")
    return slug or "GENERAL"


def gen_ref_no(db: Session, department_name: Optional[str]) -> str:
    """Reference number format: KTC-<DEPARTMENT NAME>-YY-MM-DD-001, 002, ...

    The sequence number is per department, per day, zero-padded to 3
    digits, and restarts at 001 the next calendar day (or for a different
    department). Collisions are re-checked in a loop so concurrent saves
    on the same department/day never collide.
    """
    now = dt.datetime.utcnow()
    yy = now.strftime("%y")
    mm = now.strftime("%m")
    dd = now.strftime("%d")
    dept_slug = _dept_ref_slug(department_name)
    prefix = f"KTC-{dept_slug}-{yy}-{mm}-{dd}-"
    existing = (
        db.query(Requisition.ref_no)
        .filter(Requisition.ref_no.like(f"{prefix}%"))
        .count()
    )
    seq = existing + 1
    ref_no = f"{prefix}{seq:03d}"
    while db.query(Requisition).filter(Requisition.ref_no == ref_no).first():
        seq += 1
        ref_no = f"{prefix}{seq:03d}"
    return ref_no


def _parse_requisition_line_items(raw: Optional[str]) -> list:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, ValueError):
        return []


def _parse_voucher_data(raw: Optional[str]) -> dict:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except (TypeError, ValueError):
        return {}


def requisition_to_dict(r: Requisition) -> dict:
    return {
        "id": r.id,
        "ref_no": r.ref_no,
        "requester_id": r.requester_id,
        # Prefer the free-text "Requisitioner (Full names)" typed on the
        # form itself; fall back to the account's own name for older
        # requisitions saved before this field existed.
        "requester_name": r.requester_name or (r.requester.full_name if r.requester else None),
        "requester_position": r.requester_position or (r.requester.position if r.requester else None),
        "requester_mobile": r.requester_mobile or (r.requester.telephone if r.requester else None),
        "requester_account_name": r.requester.full_name if r.requester else None,
        "requester_email": r.requester.email if r.requester else None,
        "requester_role": r.requester.role if r.requester else None,
        "requester_signature_url": r.requester.signature_url if r.requester else None,
        "department_id": r.department_id,
        "department_name": _dept_label(r.department),
        "budget_code_id": r.budget_code_id,
        "budget_code": r.budget_code.code if r.budget_code else r.budget_output_code_text,
        "budget_output": r.budget_code.output_description if r.budget_code else None,
        "activity_budget_limit": r.budget_code.allocated_amount if r.budget_code else None,
        "activity_budget_balance": r.budget_code.available_balance if r.budget_code else None,
        "activity_id": r.activity_id,
        "activity_name": r.activity.name if r.activity else None,
        "activity_details": r.activity_details,
        "subject": r.subject,
        "financial_year": r.financial_year,
        "quarter": r.quarter,
        "payment_voucher_number": r.payment_voucher_number,
        "line_items": _parse_requisition_line_items(r.line_items),
        "voucher_data": _parse_voucher_data(r.voucher_data),
        "amount_requested": r.amount_requested,
        "status": r.status,
        "current_stage": r.current_stage,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        "approvals": [
            {
                "stage": a.stage, "actor": a.actor.full_name if a.actor else None,
                "action": a.action, "comments": a.comments,
                "actor_signature_url": a.actor.signature_url if a.actor else None,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            } for a in r.approvals
        ],
        "documents": [
            {"id": d.id, "filename": d.filename, "doc_type": d.doc_type, "url": f"/files/{d.stored_path}"}
            for d in r.documents
        ],
        "accountability": {
            "status": r.accountability.status if r.accountability else None,
            "remarks": r.accountability.remarks if r.accountability else None,
            "has_voucher": any(d.doc_type == "voucher" for d in r.documents),
            "has_documents": len(r.documents) > 0,
        } if r.accountability else None,
    }


@app.get("/api/requisitions")
def list_requisitions(status_filter: Optional[str] = Query(None, alias="status"),
                       mine: bool = False,
                       db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    q = db.query(Requisition)
    if user.role == "staff" or mine:
        q = q.filter(Requisition.requester_id == user.id)
    elif user.role == "hod":
        q = q.filter(Requisition.department_id == user.department_id)
    if status_filter:
        q = q.filter(Requisition.status == status_filter)
    reqs = q.order_by(Requisition.created_at.desc()).all()
    return [requisition_to_dict(r) for r in reqs]


@app.get("/api/requisitions/{req_id}")
def get_requisition(req_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")
    return requisition_to_dict(r)


@app.post("/api/requisitions")
def create_requisition(payload: RequisitionIn, submit: bool = False,
                        db: Session = Depends(get_db),
                        user: User = Depends(require_roles("staff", "hod", "admin"))):
    bc = None
    if payload.budget_code_id:
        bc = db.query(BudgetCode).filter(BudgetCode.id == payload.budget_code_id).first()
        if not bc:
            raise HTTPException(status_code=400, detail="Selected budget code does not exist")

    dept_id = payload.department_id or user.department_id or (bc.department_id if bc else None)
    if not dept_id:
        raise HTTPException(status_code=400, detail="Please select a Department")
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=400, detail="Selected department does not exist")

    if not payload.line_items:
        raise HTTPException(status_code=400, detail="Please add at least one line item")

    total = sum((li.amount or 0) for li in payload.line_items)
    if total <= 0:
        raise HTTPException(status_code=400, detail="Please add at least one priced line item")

    r = Requisition(
        ref_no=gen_ref_no(db, dept.name),
        requester_id=user.id,
        department_id=dept.id,
        budget_code_id=payload.budget_code_id,
        activity_id=payload.activity_id,
        subject=payload.subject,
        financial_year=(payload.financial_year or "").strip() or None,
        quarter=(payload.quarter or "").strip() or None,
        requester_name=(payload.requester_name or "").strip() or user.full_name,
        requester_position=(payload.requester_position or "").strip() or user.position,
        requester_mobile=(payload.requester_mobile or "").strip() or user.telephone,
        budget_output_code_text=(payload.budget_output_code_text or "").strip() or (bc.code if bc else None),
        payment_voucher_number=(payload.payment_voucher_number or "").strip() or None,
        line_items=json.dumps([li.dict() for li in payload.line_items]),
        voucher_data=json.dumps(payload.voucher.dict()) if payload.voucher else None,
        amount_requested=total,
        status="draft",
        current_stage="hod",
    )
    db.add(r)
    db.commit()
    db.refresh(r)

    if submit:
        _submit_requisition(r, db, user)

    log_action(db, user.id, "requisition.create", r.ref_no)
    _invalidate_budget_code_caches()
    return requisition_to_dict(r)


def _submit_requisition(r: Requisition, db: Session, user: User):
    bc = r.budget_code
    if bc:
        if r.activity_id:
            act = db.query(Activity).filter(Activity.id == r.activity_id, Activity.budget_code_id == bc.id).first()
            if not act:
                raise HTTPException(status_code=400, detail="The selected activity is not part of the approved work plan for this budget code")
        if bc.available_balance < r.amount_requested:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient available budget. Available balance is UGX {bc.available_balance:,.0f}, requested UGX {r.amount_requested:,.0f}"
            )
    r.status = "submitted"
    r.current_stage = "hod"
    r.updated_at = dt.datetime.utcnow()
    db.commit()
    notify_role(db, "hod", f"New requisition {r.ref_no} awaiting your approval", "approval_request", r.id)
    _invalidate_budget_code_caches()


@app.post("/api/requisitions/{req_id}/submit")
def submit_requisition(req_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")
    if r.requester_id != user.id and user.role != "admin":
        raise HTTPException(status_code=403, detail="You can only submit your own requisitions")
    if r.status not in ("draft", "returned"):
        raise HTTPException(status_code=400, detail="Only draft or returned requisitions can be submitted")
    _submit_requisition(r, db, user)
    log_action(db, user.id, "requisition.submit", r.ref_no)
    return requisition_to_dict(r)


@app.patch("/api/requisitions/{req_id}")
def update_requisition(req_id: int, payload: RequisitionIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")
    if r.requester_id != user.id and user.role != "admin":
        raise HTTPException(status_code=403, detail="You can only edit your own requisitions")
    if r.status not in ("draft", "returned"):
        raise HTTPException(status_code=400, detail="Only draft or returned requisitions can be edited")

    bc = None
    if payload.budget_code_id:
        bc = db.query(BudgetCode).filter(BudgetCode.id == payload.budget_code_id).first()
        if not bc:
            raise HTTPException(status_code=400, detail="Selected budget code does not exist")
    if not payload.line_items:
        raise HTTPException(status_code=400, detail="Please add at least one line item")
    total = sum((li.amount or 0) for li in payload.line_items)
    if total <= 0:
        raise HTTPException(status_code=400, detail="Please add at least one priced line item")

    if payload.department_id:
        dept = db.query(Department).filter(Department.id == payload.department_id).first()
        if not dept:
            raise HTTPException(status_code=400, detail="Selected department does not exist")
        r.department_id = dept.id

    r.budget_code_id = payload.budget_code_id
    r.activity_id = payload.activity_id
    r.subject = payload.subject
    r.financial_year = (payload.financial_year or "").strip() or None
    r.quarter = (payload.quarter or "").strip() or None
    r.requester_name = (payload.requester_name or "").strip() or r.requester_name
    r.requester_position = (payload.requester_position or "").strip() or r.requester_position
    r.requester_mobile = (payload.requester_mobile or "").strip() or r.requester_mobile
    r.budget_output_code_text = (payload.budget_output_code_text or "").strip() or (bc.code if bc else r.budget_output_code_text)
    r.payment_voucher_number = (payload.payment_voucher_number or "").strip() or None
    r.line_items = json.dumps([li.dict() for li in payload.line_items])
    r.voucher_data = json.dumps(payload.voucher.dict()) if payload.voucher else None
    r.amount_requested = total
    r.updated_at = dt.datetime.utcnow()
    db.commit()
    db.refresh(r)
    log_action(db, user.id, "requisition.update", r.ref_no)
    _invalidate_budget_code_caches()
    return requisition_to_dict(r)


@app.delete("/api/requisitions/{req_id}")
def delete_requisition(req_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")
    if r.requester_id != user.id and user.role != "admin":
        raise HTTPException(status_code=403, detail="You can only delete your own requisitions")
    if r.status not in ("draft", "returned"):
        raise HTTPException(status_code=400, detail="Only draft or returned requisitions can be deleted")
    db.query(Document).filter(Document.requisition_id == req_id).delete()
    db.query(ApprovalHistory).filter(ApprovalHistory.requisition_id == req_id).delete()
    db.delete(r)
    db.commit()
    log_action(db, user.id, "requisition.delete", r.ref_no)
    _invalidate_budget_code_caches()
    return {"ok": True}


STAGE_ROLE = {"hod": "hod", "treasurer": "treasurer", "clerk": "clerk"}
NEXT_STAGE = {"hod": "treasurer", "treasurer": "clerk", "clerk": "done"}
STAGE_STATUS = {"hod": "hod_approved", "treasurer": "treasurer_approved", "clerk": "approved"}


@app.post("/api/requisitions/{req_id}/approve-action")
def approval_action(req_id: int, payload: ApprovalActionIn,
                     db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")

    stage = r.current_stage
    if stage not in STAGE_ROLE:
        raise HTTPException(status_code=400, detail="This requisition is not awaiting approval")
    if user.role != STAGE_ROLE[stage] and user.role != "admin":
        raise HTTPException(status_code=403, detail=f"Only the {STAGE_ROLE[stage].upper()} can act on this stage")
    if payload.action not in ("approve", "reject", "return"):
        raise HTTPException(status_code=400, detail="Action must be approve, reject or return")

    history = ApprovalHistory(
        requisition_id=r.id, stage=stage, actor_id=user.id,
        action=payload.action, comments=payload.comments,
    )
    db.add(history)

    if payload.action == "approve":
        if stage == "treasurer" and r.budget_code and r.budget_code.available_balance < 0:
            raise HTTPException(status_code=400, detail="Budget has since been exhausted for this code")
        r.status = STAGE_STATUS[stage]
        nxt = NEXT_STAGE[stage]
        if nxt == "done":
            r.current_stage = "done"
            r.status = "approved"
            acc = AccountabilityRecord(requisition_id=r.id, status="pending")
            db.add(acc)
            notify_role(db, "auditor", f"Requisition {r.ref_no} approved - accountability documents required", "accountability_pending", r.id)
            notify(
                db, r.requester_id,
                f"Your requisition {r.ref_no} has been fully approved. Please upload the accountability "
                f"documents (payment voucher, receipts, attendance sheets, etc.) for this requisition.",
                "accountability_pending", r.id
            )
        else:
            r.current_stage = nxt
            notify_role(db, nxt, f"Requisition {r.ref_no} awaiting your approval", "approval_request", r.id)
    elif payload.action == "reject":
        r.status = "rejected"
        r.current_stage = "closed"
        notify(db, r.requester_id, f"Your requisition {r.ref_no} was rejected", "rejection", r.id)
    else:  # return
        r.status = "returned"
        r.current_stage = "hod"
        notify(db, r.requester_id, f"Your requisition {r.ref_no} was returned for correction", "rejection", r.id)

    r.updated_at = dt.datetime.utcnow()
    db.commit()
    log_action(db, user.id, f"requisition.{payload.action}", f"{r.ref_no} at stage {stage}")
    _invalidate_budget_code_caches()
    return requisition_to_dict(r)


@app.get("/api/approvals/pending")
def pending_approvals(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if user.role not in ("hod", "treasurer", "clerk", "admin"):
        raise HTTPException(status_code=403, detail="Not an approver role")
    stage_for_role = {"hod": "hod", "treasurer": "treasurer", "clerk": "clerk"}
    if user.role == "admin":
        reqs = db.query(Requisition).filter(Requisition.current_stage.in_(["hod", "treasurer", "clerk"])).all()
    else:
        stage = stage_for_role[user.role]
        q = db.query(Requisition).filter(Requisition.current_stage == stage)
        if user.role == "hod":
            q = q.filter(Requisition.department_id == user.department_id)
        reqs = q.all()
    return [requisition_to_dict(r) for r in reqs]


# ---------------------------- Documents (Backblaze B2) -----------------------

@app.post("/api/requisitions/{req_id}/documents")
async def upload_document(req_id: int, doc_type: str = "supporting", file: UploadFile = File(...),
                           db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")

    if not r.accountability:
        raise HTTPException(
            status_code=400,
            detail="This requisition is not yet awaiting accountability documents. "
                   "It must be fully approved before documents can be uploaded."
        )
    if user.id != r.requester_id and user.role != "admin":
        raise HTTPException(
            status_code=403,
            detail="Only the original requester or an administrator can upload accountability documents for this requisition"
        )
    if r.accountability.status == "verified":
        raise HTTPException(status_code=400, detail="This requisition's accountability has already been verified")

    allowed_ext = {".pdf", ".docx", ".jpg", ".jpeg", ".png"}
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail="Only PDF, DOCX, JPG and PNG files are allowed")

    object_key = await upload_document_to_b2(file, r.id)

    doc = Document(requisition_id=r.id, filename=file.filename, stored_path=object_key,
                    doc_type=doc_type, uploaded_by=user.id)
    db.add(doc)

    if r.accountability.status == "flagged":
        r.accountability.status = "pending"

    db.commit()
    log_action(db, user.id, "document.upload", f"{file.filename} on {r.ref_no}")
    notify_role(db, "auditor", f"New accountability document uploaded for {r.ref_no}", "accountability_pending", r.id)
    return {"id": doc.id, "filename": doc.filename, "doc_type": doc.doc_type, "url": f"/files/{object_key}"}


@app.get("/files/{filename:path}")
async def stream_document(filename: str, request: Request, db: Session = Depends(get_db)):
    clean_key = filename.split("?")[0].strip("/")
    if not clean_key:
        raise HTTPException(status_code=400, detail="Filename is required")

    try:
        head = b2_client.head_object(Bucket=B2_BUCKET_NAME, Key=clean_key)
    except ClientError:
        raise HTTPException(status_code=404, detail="File not found in storage")

    file_size = head["ContentLength"]
    content_type = head.get("ContentType") or "application/octet-stream"

    range_header = request.headers.get("Range")
    get_kwargs = {"Bucket": B2_BUCKET_NAME, "Key": clean_key}
    status_code = 200
    content_range = None
    content_length = file_size

    if range_header:
        try:
            _, range_value = range_header.split("=", 1)
            start_str, end_str = range_value.split("-", 1)
            start = int(start_str) if start_str.strip() else 0
            end = int(end_str) if end_str.strip() else file_size - 1
            end = min(end, file_size - 1)
            if start > end or start >= file_size:
                raise HTTPException(status_code=416, detail="Range Not Satisfiable")
            get_kwargs["Range"] = f"bytes={start}-{end}"
            content_length = end - start + 1
            status_code = 206
            content_range = f"bytes {start}-{end}/{file_size}"
        except (ValueError, AttributeError):
            pass

    try:
        b2_response = b2_client.get_object(**get_kwargs)
    except ClientError:
        raise HTTPException(status_code=502, detail="Storage fetch error")

    doc = db.query(Document).filter(Document.stored_path == clean_key).first()
    if doc:
        display_name = doc.filename
    else:
        display_name = clean_key.split("/")[-1]

    def _stream(body, chunk_size: int = 65536):
        with body as stream:
            while True:
                chunk = stream.read(chunk_size)
                if not chunk:
                    break
                yield chunk

    headers = {
        "Content-Type": content_type,
        "Content-Length": str(content_length),
        "Accept-Ranges": "bytes",
        "Content-Disposition": f'inline; filename="{display_name}"',
        "Cache-Control": "private, max-age=3600",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, HEAD, OPTIONS",
        "Access-Control-Expose-Headers": "Content-Range, Content-Length, Accept-Ranges",
    }
    if content_range:
        headers["Content-Range"] = content_range

    return StreamingResponse(
        _stream(b2_response["Body"]),
        status_code=status_code,
        headers=headers,
        media_type=content_type,
    )


# ---------------------------- Accountability --------------------------------

@app.post("/api/requisitions/{req_id}/accountability")
def update_accountability(req_id: int, payload: AccountabilityIn,
                           db: Session = Depends(get_db), user: User = Depends(require_roles("auditor", "admin"))):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r or not r.accountability:
        raise HTTPException(status_code=404, detail="No accountability record found for this requisition")

    if payload.status == "verified":
        if not r.documents:
            raise HTTPException(
                status_code=400,
                detail="Cannot verify: no accountability documents have been uploaded for this requisition yet."
            )
        has_voucher = any(d.doc_type == "voucher" for d in r.documents)
        if not has_voucher:
            raise HTTPException(
                status_code=400,
                detail="Cannot verify: a Payment Voucher must be uploaded for this requisition before it can be verified."
            )

    r.accountability.status = payload.status
    r.accountability.remarks = payload.remarks
    r.accountability.auditor_id = user.id
    r.accountability.updated_at = dt.datetime.utcnow()
    if payload.status == "verified":
        r.status = "accounted"
    db.commit()
    log_action(db, user.id, "accountability.update", f"{r.ref_no} -> {payload.status}")
    if payload.status == "flagged":
        notify(
            db, r.requester_id,
            f"Accountability documents for {r.ref_no} were flagged: {payload.remarks or 'see remarks'}",
            "rejection", r.id
        )
    elif payload.status == "verified":
        notify(db, r.requester_id, f"Accountability for {r.ref_no} has been verified", "approval_completed", r.id)
    return requisition_to_dict(r)


@app.get("/api/accountability/pending")
def accountability_pending(db: Session = Depends(get_db), user: User = Depends(require_roles("auditor", "admin"))):
    reqs = db.query(Requisition).join(AccountabilityRecord).filter(AccountabilityRecord.status != "verified").all()
    return [requisition_to_dict(r) for r in reqs]


# ---------------------------- Notifications ---------------------------------

@app.get("/api/notifications")
def list_notifications(unread_only: bool = False, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    q = db.query(Notification).filter(Notification.user_id == user.id)
    if unread_only:
        q = q.filter(Notification.is_read == False)
    items = q.order_by(Notification.created_at.desc()).limit(50).all()
    return [
        {"id": n.id, "message": n.message, "category": n.category, "is_read": n.is_read,
         "created_at": n.created_at.isoformat(), "requisition_id": n.link_requisition_id}
        for n in items
    ]


@app.patch("/api/notifications/{note_id}/read")
def mark_notification_read(note_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    n = db.query(Notification).filter(Notification.id == note_id, Notification.user_id == user.id).first()
    if not n:
        raise HTTPException(status_code=404, detail="Notification not found")
    n.is_read = True
    db.commit()
    return {"ok": True}


@app.patch("/api/notifications/read-all")
def mark_all_read(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    db.query(Notification).filter(Notification.user_id == user.id, Notification.is_read == False).update({"is_read": True})
    db.commit()
    return {"ok": True}


# ---------------------------- Work Plan PDF Report ---------------------------
# Static narrative text carried over from the Council's Annual Work Plan &
# Budget document (forward, Town Clerk message, executive summary). The two
# data tables below them (Revenue Sources, Annual Workplan Budget) are NOT
# static — they're rendered live from the database at export time.

REPORT_FORWARD_PARAS = [
    "It is my pleasure to present the Karugutu Town Council Annual Work Plan and Budget for the Financial Year 2026/2027, which provides the Council's implementation roadmap for delivering quality public services and advancing sustainable socio-economic transformation within our municipality. This Work Plan and Budget operationalizes the Karugutu Town Council Five-Year Strategic Development Plan (2025–2030) by translating its strategic priorities into annual programmes, budget outputs, measurable performance targets, and resource allocations that respond to the aspirations and needs of our people.",
    "The FY 2026/27 Work Plan has been prepared in accordance with the Constitution of the Republic of Uganda, the Local Governments Act, the Public Finance Management Act, the Programme-Based Budgeting Framework, the Fourth National Development Plan (NDP IV), and other relevant Government policies and guidelines. It reflects our commitment to prudent financial management, transparency, accountability, citizen participation, and results-oriented service delivery.",
    "During the financial year, the Council will prioritize investments in road infrastructure maintenance, engineering services, physical planning, environmental management, local revenue enhancement, financial accountability, institutional capacity building, agricultural production, community empowerment, and improved governance. These interventions are intended to create an enabling environment for economic growth, improve public service delivery, promote orderly urban development, and enhance the quality of life of the residents of Karugutu Town Council.",
    "The successful implementation of this Annual Work Plan will depend on the collective efforts of all stakeholders, including the Central Government, Ntoroko District Local Government, Development Partners, the Technical Staff, Council Members, the Private Sector, Civil Society Organizations, Community-Based Organizations, cultural and religious leaders, and the people of Karugutu Town Council. I therefore call upon every stakeholder to actively participate in the implementation, monitoring, and evaluation of the planned interventions to ensure that the intended development outcomes are achieved.",
    "On behalf of the Council, I extend my sincere appreciation to the Government of Uganda, the Ministry of Local Government, the Ministry of Finance, Planning and Economic Development, Ntoroko District Local Government, our Development Partners, and all stakeholders whose technical and financial support has contributed to the preparation of this Annual Work Plan and Budget.",
    "I am confident that, through prudent resource management, strong partnerships, and unwavering commitment, Karugutu Town Council will continue to make significant progress towards achieving its strategic vision of a well-planned, prosperous, resilient, and inclusive urban community.",
]
REPORT_FORWARD_QUOTE = "\u201cTogether, we shall build a stronger, more prosperous, and sustainable Karugutu Town Council.\u201d"
REPORT_FORWARD_SIGNOFF = ["Hon Maate Raphael G", "THE CHAIRPERSON LC III", "Karugutu Town Council", "Ntoroko District Local Government"]

REPORT_CLERK_INTRO = [
    "It is my privilege to present the Karugutu Town Council Annual Work Plan and Budget for the 2026/2027 Financial Year. This document serves as our operational blueprint. It translates our overarching Five-Year Strategic Development Plan (2025–2030) into actionable projects, clear performance targets, and firm financial commitments.",
    "We developed this work plan in strict compliance with Uganda\u2019s legal frameworks. These include the Constitution, the Local Governments Act, the Public Finance Management Act, and the Fourth National Development Plan (NDP IV). It anchors our operations in Programme-Based Budgeting to guarantee transparency, strict accountability, and high-quality service delivery.",
    "For FY 2026/2027, we link every single shilling allocated directly to measurable targets, timelines, and specific oversight officers. This results-oriented setup ensures straightforward monitoring and guarantees maximum value for public funds.",
    "Our key structural priorities for this financial year focus on:",
]
REPORT_CLERK_BULLETS = [
    "Infrastructure: Expanding road networks and engineering projects.",
    "Urban Growth: Enhancing physical planning and orderly development.",
    "Finance: Boosting local revenue and enforcing strict financial controls.",
    "Community: Improving public health sanitation and reinforcing local governance.",
    "Economy: Driving agricultural production and local economic initiatives.",
]
REPORT_CLERK_OUTRO = [
    "Achieving these milestones demands deep collaboration. We rely on teamwork across all political leaders, technical staff, Ntoroko District Local Government, ministries, and civil society. As Accounting Officer, I will strictly enforce compliance with all national laws, financial regulations, and procurement standards.",
    "I extend my gratitude to the Mayor, the Executive Committee, our Technical Planning Committee, and our development partners. Your valuable input made this comprehensive plan possible.",
    "As implementation begins, I urge all council staff and stakeholders to champion professionalism and integrity. Let us work together to build a prosperous, resilient, and well-governed Karugutu Town Council.",
]
REPORT_CLERK_SIGNOFF = ["TOWN CLERK", "Karugutu Town Council", "Ntoroko District Local Government"]

# Executive summary: list of ("h3"|"h4"|"p"|"bullets", content) tuples, rendered in order.
REPORT_EXEC_SUMMARY = [
    ("h3", "Karugutu Town Council Revenue Analysis for FY 2026/27"),
    ("p", "The Council\u2019s approved revenue framework for FY 2026/27 amounts to UGX 327,113,640, reflecting a balanced financing strategy that combines Central Government Transfers, Locally Raised Revenue (LRR), and limited development partner support. Central Government Transfers constitute the largest share of the resource envelope at UGX 175,991,640 (53.8%), primarily comprising Urban Unconditional Grant (Non-Wage), Urban Discretionary Development Equalization Grant (DDEG), and the Uganda Road Fund, thereby providing the financial foundation for recurrent service delivery and infrastructure maintenance."),
    ("p", "Locally Raised Revenue is projected at UGX 151,120,000 (46.2%), demonstrating the Council's commitment to strengthening fiscal autonomy through enhanced revenue mobilization. The principal local revenue streams include Market Revenue (UGX 80.0 million), Council Property and Asset Revenues (UGX 19.4 million), Business Licensing and Trade Regulation (UGX 12.1 million), and Physical Planning and Development Control (UGX 12.47 million). Additional revenues are expected from property rates, public health services, transport and parking fees, environmental charges, advertising, enforcement penalties, and other administrative fees, thereby diversifying the local revenue base and reducing dependence on intergovernmental transfers."),
    ("p", "Overall, the Council\u2019s FY 2026/27 revenue structure reflects a strategic emphasis on improving local revenue performance while leveraging Government transfers to finance priority programmes, enhance service delivery, strengthen urban governance, and support the sustainable socio-economic development of Karugutu Town Council."),
    ("h3", "Annual Work Plan and Budget Performance Framework Analysis for FY 2026/27"),
    ("p", "On the other hand, the Council\u2019s Annual Work Plan and Budget for FY 2026/27 operationalizes the Karugutu Town Council Five-Year Strategic Development Plan (2025\u20132030) by translating its strategic objectives, priority programmes, and investment interventions into measurable annual outputs, activities, performance targets, and budget allocations. The Work Plan aligns with the Programme-Based Budgeting (PBB) Framework, the Fourth National Development Plan (NDP IV), the Sustainable Development Goals (SDGs), and other relevant national policy and legal frameworks."),
    ("p", "The FY 2026/27 Work Plan serves as the first annual implementation instrument under the Strategic Development Plan, focusing on strengthening institutional governance, expanding infrastructure development, improving urban planning and environmental management, enhancing local revenue mobilisation, promoting local economic development, and improving access to quality public services."),
    ("h4", "1. Budget Allocation and Strategic Investment Priorities"),
    ("p", "The Council has strategically allocated approximately UGX 327.1 million across eight functional departments to finance the 2026/27 Annual Work Plan and a few priority interventions identified in the Five-Year Strategic Development Plan (2025\u20132030)."),
    ("p", "The 2026/27 Annual Work Plan and Budget composition demonstrates a strategic emphasis on infrastructure development as a catalyst for economic transformation while simultaneously investing in governance, institutional strengthening, financial accountability, social development, and environmental sustainability."),
    ("h4", "2. Financing Strategy and Budget Implementation"),
    ("p", "Implementation of the Annual Work Plan will be financed through a combination of Locally Raised Revenue, the Urban Unconditional Grant (Non-Wage), the Uganda Road Fund (URF), and the Urban Discretionary Development Equalization Grant (DDEG)."),
    ("p", "Routine administrative operations, planning, financial management, community development, governance, monitoring, and regulatory functions are primarily financed through locally generated revenue and unconditional government transfers. Capital-intensive investments are financed through conditional development grants."),
    ("p", "Quarterly budget phasing has been adopted to promote efficient cash flow management, timely implementation of activities, fiscal discipline, and continuous service delivery throughout the financial year."),
    ("h4", "3. Results-Based Performance Management Framework"),
    ("p", "The FY 2026/27 Annual Work Plan adopts a Results-Based Management (RBM) approach in which every budget output is linked to clearly defined performance indicators, baseline values, annual targets, quarterly milestones, funding sources, and responsible officers."),
    ("p", "Priority performance interventions include:"),
    ("bullets", [
        "Improving road infrastructure and engineering services to enhance connectivity and economic productivity.",
        "Strengthening local revenue mobilization and financial compliance to improve fiscal sustainability.",
        "Promoting orderly urban growth through spatial planning, GIS mapping, and effective development control.",
        "Enhancing democratic governance through Council meetings, Standing Committees, Ward Development Committees, public barazas, and citizen engagement.",
        "Supporting agricultural production, veterinary services, food safety, and enterprise development to stimulate household incomes.",
        "Expanding community empowerment programmes focusing on gender equality, youth employment, disability inclusion, child protection, and vulnerable groups.",
        "Strengthening internal controls, financial accountability, procurement compliance, and value-for-money assurance through independent internal audit functions.",
    ]),
    ("h4", "4. Monitoring, Evaluation and Accountability"),
    ("p", "Implementation of the Annual Work Plan will be monitored through quarterly performance reviews, Technical Planning Committee meetings, field supervision, internal audits, Council oversight, statutory reporting, and annual performance assessments."),
    ("p", "The integration of digital planning tools, electronic records management, GIS-based planning systems, and performance reporting further reinforces the Council's commitment to transparency, innovation, and effective public sector management."),
    ("h4", "5. Strategic Outlook"),
    ("p", "The FY 2026/27 Annual Work Plan represents the practical implementation roadmap for the Karugutu Town Council Five-Year Strategic Development Plan (2025\u20132030), providing a clear pathway for achieving the Council's long-term development aspirations while ensuring accountability, efficiency, and value for public resources."),
    ("h3", "Annual Budget Framework for FY 2026/27"),
    ("p", "The Annual Budget Framework for FY 2026/27 provides the strategic financial planning and resource allocation mechanism through which Karugutu Town Council will implement its development priorities and statutory mandates during the financial year 1 July 2026 to 30 June 2027."),
    ("p", "The budget framework is financed through a combination of locally generated revenue, central government transfers, external financing where applicable, and other lawful sources of revenue, allocated across programmes, departments, and budget outputs."),
    ("p", "The framework adopts a results-based and programme-oriented approach, ensuring that financial resources are directly linked to measurable outputs, outcomes, and performance indicators, guided by principles of fiscal discipline, value for money, equity, transparency, and accountability."),
    ("p", "To enhance accountability and effective budget execution, the Council will undertake participatory planning, stakeholder consultations, budget conferences, technical planning committee reviews, council approvals, and regular monitoring and evaluation."),
    ("h3", "Annual Workplan and Budget Output Performance Targets for FY 2026/27"),
    ("p", "The Karugutu Town Council Annual Workplan and Budget Output Performance Targets document serves as the formal ex-ante planning instrument, translating national strategic objectives into localized service delivery milestones, formulated under the Programme Budgeting System (PBS) and the Programme Implementation Action Plans (PIAP) framework."),
    ("p", "Structured hierarchically by Vote, Programme, Sub-Sub Programme, and distinct Budget Outputs, this workplan maps out every operational shilling against verifiable targets."),
    ("p", "Key Features of the Annual Workplan Framework:"),
    ("bullets", [
        "Logical Performance Tracking: Every expenditure line item links directly to a designated baseline value, a clear quarterly cash distribution projection, and a specified PIAP Output Indicator.",
        "Decentralized Service Monitoring: Operational target structures are embedded directly into daily service points.",
        "Verifiable Statutory Accountability: The planned targets establish a strict compliance baseline for ex-post audits, used during Annual Local Government Performance Assessment (LGPA) cycles.",
    ]),
    ("p", "Ultimately, the detailed document hereunder stands as Karugutu Town Council's binding operational commitment to transparency, evidence-based public financial management, and equitable development."),
]

# Paragraph/table styles used only by the PDF report.
_pdf_styles = getSampleStyleSheet()
_pdf_body = ParagraphStyle("RptBody", parent=_pdf_styles["Normal"], fontSize=9.3, leading=13.5, alignment=TA_JUSTIFY, spaceAfter=7)
_pdf_bullet = ParagraphStyle("RptBullet", parent=_pdf_body, leftIndent=14, spaceAfter=4)
_pdf_h1 = ParagraphStyle("RptH1", parent=_pdf_styles["Heading1"], fontSize=14, textColor=colors.HexColor("#146B5F"), spaceBefore=4, spaceAfter=10)
_pdf_h2 = ParagraphStyle("RptH2", parent=_pdf_styles["Heading2"], fontSize=12.5, textColor=colors.HexColor("#0A1F2B"), spaceBefore=6, spaceAfter=8)
_pdf_h3 = ParagraphStyle("RptH3", parent=_pdf_styles["Heading3"], fontSize=11, textColor=colors.HexColor("#146B5F"), spaceBefore=10, spaceAfter=5)
_pdf_h4 = ParagraphStyle("RptH4", parent=_pdf_styles["Heading4"], fontSize=10, textColor=colors.HexColor("#0A1F2B"), spaceBefore=6, spaceAfter=3)
_pdf_signoff = ParagraphStyle("RptSignoff", parent=_pdf_body, alignment=TA_CENTER, spaceAfter=2)
_pdf_cover_title = ParagraphStyle("RptCoverTitle", parent=_pdf_styles["Title"], fontSize=22, alignment=TA_CENTER, spaceAfter=6)
_pdf_cover_sub = ParagraphStyle("RptCoverSub", parent=_pdf_styles["Normal"], fontSize=15, alignment=TA_CENTER, textColor=colors.HexColor("#146B5F"), spaceAfter=4)
_pdf_cover_line = ParagraphStyle("RptCoverLine", parent=_pdf_styles["Normal"], fontSize=10.5, alignment=TA_CENTER, spaceAfter=3)
_pdf_cell = ParagraphStyle("RptCell", parent=_pdf_styles["Normal"], fontSize=6.6, leading=8.2, wordWrap="CJK")
_pdf_cell_b = ParagraphStyle("RptCellB", parent=_pdf_cell, fontName="Helvetica-Bold")
_pdf_cell_r = ParagraphStyle("RptCellR", parent=_pdf_cell, alignment=TA_RIGHT)
_pdf_cell_rb = ParagraphStyle("RptCellRB", parent=_pdf_cell_b, alignment=TA_RIGHT)

# The letterhead banner (Uganda coat of arms + "Karugutu Town Council" wordmark
# + Ntoroko District Local Government seal) used on the report's cover page,
# matching the letterhead already used at the top of the Word (.docx) version
# of this report. Expected to live at assets/ktc_letterhead.png alongside
# this file; the PDF still builds fine (just without the banner) if it's
# ever missing, so a missing asset never breaks report generation.
_PDF_LETTERHEAD_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "ktc_letterhead.png")


def _pdf_letterhead_image(max_width=460):
    """Centered Image flowable of the letterhead banner, scaled to max_width
    pt while preserving its native aspect ratio (so both logos keep their
    proportions exactly as in the .docx letterhead). Returns None if the
    asset file isn't present."""
    if not os.path.exists(_PDF_LETTERHEAD_PATH):
        return None
    try:
        from PIL import Image as PILImage
        with PILImage.open(_PDF_LETTERHEAD_PATH) as im:
            src_w, src_h = im.size
    except Exception:
        logger.warning("Could not read letterhead image at %s", _PDF_LETTERHEAD_PATH)
        return None
    img = RLImage(_PDF_LETTERHEAD_PATH, width=max_width, height=max_width * (src_h / src_w))
    img.hAlign = "CENTER"
    return img


def _pdf_money(n) -> str:
    """Format a number the same way the frontend's money() helper does."""
    return f"{parse_amount(n):,.0f}"


def _pdf_narrative_flowables(paras, quote=None, signoff=None, bullets=None):
    """Build a forward/message page: intro paragraphs, optional bullets, optional quote + signature block."""
    flow = [Paragraph(t, _pdf_body) for t in paras]
    if bullets:
        flow += [Paragraph("\u2022 " + b, _pdf_bullet) for b in bullets]
    if quote:
        flow.append(Spacer(1, 6))
        flow.append(Paragraph(f"<i>{quote}</i>", ParagraphStyle("Q", parent=_pdf_body, alignment=TA_CENTER)))
    if signoff:
        flow.append(Spacer(1, 24))
        for i, line in enumerate(signoff):
            style = _pdf_signoff if i > 0 else ParagraphStyle("QSig", parent=_pdf_signoff, fontName="Helvetica-Bold")
            flow.append(Paragraph(line, style))
    return flow


def _pdf_exec_summary_flowables():
    flow = []
    for kind, content in REPORT_EXEC_SUMMARY:
        if kind == "h3":
            flow.append(Paragraph(content, _pdf_h3))
        elif kind == "h4":
            flow.append(Paragraph(content, _pdf_h4))
        elif kind == "p":
            flow.append(Paragraph(content, _pdf_body))
        elif kind == "bullets":
            flow += [Paragraph("\u2022 " + b, _pdf_bullet) for b in content]
    return flow


def _pdf_dept_summary_table(codes, committed_map):
    """Mirrors the frontend's renderDeptSummaryTable() grouping — one row per department.

    Labels each row with "<code>: <name>" (via _dept_code_and_name), matching
    the department dropdowns and this same table on screen."""
    grouped = {}
    for c in codes:
        name = _dept_code_and_name(c.department) or "Unassigned"
        row = grouped.setdefault(name, {"q1": 0.0, "q2": 0.0, "q3": 0.0, "q4": 0.0, "total": 0.0, "uncommitted": 0.0})
        q1, q2, q3, q4 = parse_amount(c.q1_amount), parse_amount(c.q2_amount), parse_amount(c.q3_amount), parse_amount(c.q4_amount)
        total = q1 + q2 + q3 + q4
        row["q1"] += q1; row["q2"] += q2; row["q3"] += q3; row["q4"] += q4
        row["total"] += total
        row["uncommitted"] += total - committed_map.get(c.id, 0.0)
    rows = sorted(grouped.items(), key=lambda kv: -kv[1]["total"])

    header = ["Department", "Q1 (UGX)", "Q2 (UGX)", "Q3 (UGX)", "Q4 (UGX)", "Total Budget (UGX)", "Uncommitted Balance (UGX)"]
    # Every cell is a Paragraph, not a plain string, so long department names
    # and header labels wrap onto extra lines inside their own cell rather
    # than spilling over the neighboring column.
    data = [[Paragraph(h, _pdf_cell_b) for h in header]]
    grand = {"q1": 0.0, "q2": 0.0, "q3": 0.0, "q4": 0.0, "total": 0.0, "uncommitted": 0.0}
    for name, v in rows:
        row_vals = [name, _pdf_money(v["q1"]), _pdf_money(v["q2"]), _pdf_money(v["q3"]), _pdf_money(v["q4"]), _pdf_money(v["total"]), _pdf_money(v["uncommitted"])]
        data.append([
            Paragraph(row_vals[0], _pdf_cell),
            *[Paragraph(x, _pdf_cell_r) for x in row_vals[1:]],
        ])
        for k in grand:
            grand[k] += v[k]
    footer_vals = ["Sub Total", _pdf_money(grand["q1"]), _pdf_money(grand["q2"]), _pdf_money(grand["q3"]), _pdf_money(grand["q4"]), _pdf_money(grand["total"]), _pdf_money(grand["uncommitted"])]
    data.append([
        Paragraph(footer_vals[0], _pdf_cell_b),
        *[Paragraph(x, _pdf_cell_rb) for x in footer_vals[1:]],
    ])

    t = Table(data, colWidths=[182, 65, 65, 65, 65, 80, 88], repeatRows=1)
    t.setStyle(_pdf_table_style(header_rows=1, footer_rows=1))
    return t


def _pdf_table_style(header_rows=1, footer_rows=0):
    # All cell content is wrapped in Paragraph flowables (see the _pdf_cell*
    # styles), so text wraps to a new line inside its own cell instead of
    # overflowing into the next column/row. FONTNAME/FONTSIZE below are a
    # fallback for any plain-string cell that might slip through; they have
    # no effect on Paragraph cells, which use their own ParagraphStyle.
    style = [
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.HexColor("#146B5F")),
        ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.white),
        ("FONTNAME", (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C7D2D0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, header_rows), (-1, -(footer_rows + 1)), [colors.white, colors.HexColor("#EAF6F1")]),
    ]
    if footer_rows:
        style.append(("BACKGROUND", (0, -footer_rows), (-1, -1), colors.HexColor("#EAF0EF")))
        style.append(("FONTNAME", (0, -footer_rows), (-1, -1), "Helvetica-Bold"))
    return TableStyle(style)


# ---------------------------------------------------------------------------
# Revenue Sources Summary — the 3 fixed PBS categories.
# Mirrors REVENUE_SUMMARY_CATEGORIES / categorizeRevenueSource() in the
# frontend (index.html) so the PDF report's Revenue Sources Summary table
# matches the in-app table exactly.
# ---------------------------------------------------------------------------
REVENUE_SUMMARY_CATEGORIES = [
    {
        "key": "gou",
        "pbs_fund_code": "001",
        "source_of_financing_name": "Central Government Transfers (GoU)",
        "functional_definition": "Regular structural wage, non-wage recurrent, and capital grants.",
    },
    {
        "key": "lr",
        "pbs_fund_code": "002",
        "source_of_financing_name": "Locally Raised Revenues (LR)",
        "functional_definition": "Internally collected fees, levies, licenses, and operational fines.",
    },
    {
        "key": "mdp",
        "pbs_fund_code": "400",
        "source_of_financing_name": "Multi-lateral Development Partners",
        "functional_definition": "International institutional donor funding (e.g., World Bank, UNICEF).",
    },
]

# Lookup used by _normalize_revenue_source_fields() (defined earlier in this
# file, near revenue_source_to_out) to resolve a saved category shorthand
# ('gou' / 'lr' / 'mdp') back to its full Council-approved label.
_REVENUE_KEY_TO_FULL = {cat["key"]: cat for cat in REVENUE_SUMMARY_CATEGORIES}


def categorize_revenue_source(r: "RevenueSource") -> Optional[str]:
    """Python port of the frontend's categorizeRevenueSource(): matches a
    revenue source to one of the 3 fixed PBS categories by fund code or
    keywords in the source name, or returns None if it fits none of them
    (such entries are excluded from the summary totals, same as the UI)."""
    code = (r.pbs_fund_code or "").strip()
    name = (r.source_of_financing_name or "").lower()

    if code == "001" or "central government" in name or "gou" in name or "transfers" in name:
        return "gou"
    if code == "002" or "locally raised" in name or "local revenue" in name or re.search(r"\blr\b", name):
        return "lr"
    if code == "400" or "multi-lateral" in name or "multilateral" in name or "development partner" in name or "donor" in name:
        return "mdp"
    return None


def _pdf_revenue_summary_table(sources):
    """Mirrors the frontend's Revenue Sources Summary — the 3 fixed PBS categories."""
    outs = [revenue_source_to_out(r) for r in sources]
    totals = {c["key"]: 0.0 for c in REVENUE_SUMMARY_CATEGORIES}
    for r, out in zip(sources, outs):
        key = categorize_revenue_source(r)
        if key:
            totals[key] += out.approved_budget_amount

    header = ["Revenue Source", "Revenue Source Definition", "Revenue Source Amount (UGX)"]
    # Paragraph-wrapped cells throughout: "Revenue Source" (PBS Fund Code +
    # Source of Financing Name combined) and "Revenue Source Definition" are
    # free-text and easily longer than their column, so they need to wrap
    # onto extra lines rather than overlap the "Revenue Source Amount"
    # column next to them.
    data = [[Paragraph(h, _pdf_cell_b) for h in header]]
    grand_total = 0.0
    for cat in REVENUE_SUMMARY_CATEGORIES:
        amt = totals[cat["key"]]
        grand_total += amt
        data.append([
            Paragraph(f'{cat["pbs_fund_code"]} — {cat["source_of_financing_name"]}', _pdf_cell),
            Paragraph(cat["functional_definition"], _pdf_cell),
            Paragraph(_pdf_money(amt), _pdf_cell_r),
        ])
    data.append([
        Paragraph("Total Revenue", _pdf_cell_b), Paragraph("", _pdf_cell_b),
        Paragraph(_pdf_money(grand_total), _pdf_cell_rb),
    ])

    t = Table(data, colWidths=[220, 225, 90], repeatRows=1)
    t.setStyle(_pdf_table_style(header_rows=1, footer_rows=1))
    return t


def _pdf_revenue_detail_table(sources):
    """The full 'Revenue Source by Category' table — one category head row per
    revenue source plus its sub rows (revenue items), matching the in-app table."""
    header = ["PBS Fund Code", "Revenue Source", "Functional Definition", "Item", "Estimate (UGX)", "Category Total (UGX)"]
    # Paragraph-wrapped cells: revenue source names, functional definitions
    # and item descriptions are all free text, so each wraps to extra lines
    # inside its own cell instead of overlapping the columns beside it.
    data = [[Paragraph(h, _pdf_cell_b) for h in header]]
    span_cmds = []
    for r in sources:
        out = revenue_source_to_out(r)
        items = out.items or []
        row_idx = len(data)
        if items:
            for i, it in enumerate(items):
                data.append([
                    Paragraph(r.pbs_fund_code or "" if i == 0 else "", _pdf_cell),
                    Paragraph(r.source_of_financing_name if i == 0 else "", _pdf_cell),
                    Paragraph(r.functional_definition or "" if i == 0 else "", _pdf_cell),
                    Paragraph(it.description, _pdf_cell),
                    Paragraph(_pdf_money(it.amount), _pdf_cell_r),
                    Paragraph(_pdf_money(out.category_total) if i == 0 else "", _pdf_cell_rb),
                ])
            span_cmds.append(("SPAN", (0, row_idx), (0, row_idx + len(items) - 1)))
            span_cmds.append(("SPAN", (1, row_idx), (1, row_idx + len(items) - 1)))
            span_cmds.append(("SPAN", (2, row_idx), (2, row_idx + len(items) - 1)))
            span_cmds.append(("SPAN", (5, row_idx), (5, row_idx + len(items) - 1)))
        else:
            data.append([
                Paragraph(r.pbs_fund_code or "", _pdf_cell), Paragraph(r.source_of_financing_name, _pdf_cell),
                Paragraph(r.functional_definition or "", _pdf_cell), Paragraph("\u2014", _pdf_cell),
                Paragraph("", _pdf_cell_r), Paragraph(_pdf_money(out.category_total), _pdf_cell_rb),
            ])

    t = Table(data, colWidths=[52, 110, 155, 130, 62, 75], repeatRows=1)
    style = _pdf_table_style(header_rows=1)
    for cmd in span_cmds:
        style.add(*cmd)
    t.setStyle(style)
    return t


def _pdf_workplan_table(codes, committed_map):
    """The full 'Annual Workplan for the FY' budget-code table — every column
    shown in the in-app table, one row per budget code."""
    header = ["Dept", "Service Area", "Programme", "Sub Programme", "Code", "Output Description", "PIAP Description",
               "PIAP Indicator", "Unit", "Baseline", "Target", "Actual", "Q1", "Q2", "Q3", "Q4", "Total", "Funding Source", "Responsible Party"]
    # Which columns hold numbers, right-aligned for readability (money/counts
    # read better right-aligned than left-aligned).
    numeric_cols = {12, 13, 14, 15, 16}
    data = [[Paragraph(h, _pdf_cell_b) for h in header]]
    for c in codes:
        # Baseline/Target: many PIAP indicators record these as narrative
        # text (e.g. "25% of Council area mapped for vectors") rather than
        # a plain figure — show the original text (baseline_note/
        # target_note) when present, since it's the actual source-of-truth
        # value; otherwise fall back to the numeric figure.
        baseline_display = c.baseline_note or _pdf_money(c.baseline_value)
        target_display = c.target_note or _pdf_money(c.planned_target)
        row = [
            c.department.name if c.department else "\u2014", c.service_area or "", c.programme or "", c.sub_programme or "",
            c.code, c.output_description or "", c.piap_output_description or "", c.piap_output_indicator or "",
            c.unit_of_measure or "", baseline_display, target_display, c.actual_output or "",
            _pdf_money(c.q1_amount), _pdf_money(c.q2_amount), _pdf_money(c.q3_amount), _pdf_money(c.q4_amount),
            _pdf_money(c.allocated_amount), c.funding_source or "", c.responsible_party or "",
        ]
        data.append([Paragraph(str(v), _pdf_cell_r if i in numeric_cols else _pdf_cell) for i, v in enumerate(row)])

    # 19 columns on a landscape A4 page (doc margins 28pt each side) leave
    # ~786pt of usable width. These widths sum to ~764pt, so the table fits
    # inside that frame with room to spare — a table wider than the frame is
    # what was causing columns to bleed into each other. Every cell is still
    # a Paragraph, so any text too long for its column wraps onto a new line
    # inside that cell instead of spilling into the next one.
    col_widths = [40, 38, 38, 38, 30, 62, 54, 54, 26, 24, 24, 28, 44, 44, 44, 44, 48, 38, 42]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style(header_rows=1))
    return t


def build_workplan_report_pdf(wp: "WorkPlan", codes: list, committed_map: dict, sources: list) -> io.BytesIO:
    """Assembles the full Annual Work Plan & Budget PDF: static narrative
    sections plus the live Revenue Sources and Annual Workplan Budget tables."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=28, bottomMargin=28, leftMargin=28, rightMargin=28)

    letterhead = _pdf_letterhead_image()
    story = [Spacer(1, 90)]
    if letterhead is not None:
        # Banner already contains the "Karugutu Town Council" wordmark
        # between the two logos, exactly as in the .docx letterhead, so the
        # separate "Karugutu Town Council" text line isn't repeated here.
        story.append(letterhead)
        story.append(Spacer(1, 18))
        story.append(Paragraph("Ntoroko District Local Government", _pdf_cover_line))
        story.append(Paragraph("P.O. Box 568, Fort Portal, Uganda", _pdf_cover_line))
    else:
        story.append(Paragraph("Karugutu Town Council", _pdf_cover_line))
        story.append(Paragraph("Ntoroko District Local Government", _pdf_cover_line))
        story.append(Paragraph("P.O. Box 568, Fort Portal, Uganda", _pdf_cover_line))
    story += [
        Spacer(1, 40),
        Paragraph("ANNUAL WORK PLAN AND BUDGET", _pdf_cover_title),
        Paragraph(f"FY {wp.financial_year}", _pdf_cover_sub),
        Spacer(1, 40),
        Paragraph("Prepared by: Karugutu Town Council Technical Planning Committee", _pdf_cover_line),
        Paragraph(f"Generated on {dt.datetime.utcnow().strftime('%d %B %Y')}", _pdf_cover_line),
        PageBreak(),
        Paragraph("FORWARD BY CHAIRPERSON LC III KARUGUTU TOWN COUNCIL", _pdf_h1),
    ]
    story += _pdf_narrative_flowables(REPORT_FORWARD_PARAS, quote=REPORT_FORWARD_QUOTE, signoff=REPORT_FORWARD_SIGNOFF)
    story.append(PageBreak())
    story.append(Paragraph("MESSAGE FROM THE TOWN CLERK, KARUGUTU TOWN COUNCIL", _pdf_h1))
    story += _pdf_narrative_flowables(REPORT_CLERK_INTRO, bullets=REPORT_CLERK_BULLETS)
    story += _pdf_narrative_flowables(REPORT_CLERK_OUTRO, signoff=REPORT_CLERK_SIGNOFF)
    story.append(PageBreak())
    story.append(Paragraph("EXECUTIVE SUMMARY", _pdf_h1))
    story += _pdf_exec_summary_flowables()
    story.append(PageBreak())

    story.append(Paragraph("Department Budget Summary", _pdf_h2))
    story.append(_pdf_dept_summary_table(codes, committed_map))
    story.append(Spacer(1, 16))
    story.append(Paragraph("Revenue Sources Summary", _pdf_h2))
    story.append(_pdf_revenue_summary_table(sources))
    story.append(PageBreak())
    story.append(Paragraph("APPROVED COUNCIL BUDGET FRAMEWORK PAPER AND PRELIMINARY REVENUE AND EXPENDITURE ESTIMATES FOR FY 2026/2027", _pdf_h2))
    story.append(_pdf_revenue_detail_table(sources))
    story.append(PageBreak())
    story.append(Paragraph("APPROVED COUNCIL ANNUAL WORK PLAN AND BUDGET ESTIMATES FOR FY 2026/2027", _pdf_h2))
    story.append(_pdf_workplan_table(codes, committed_map))

    doc.build(story)
    buf.seek(0)
    return buf


@app.get("/api/work-plans/{wp_id}/report-pdf")
def download_workplan_report_pdf(wp_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Streams the full Annual Work Plan & Budget PDF (narrative + Revenue
    Sources + Annual Workplan Budget tables) for one work plan."""
    wp = db.query(WorkPlan).filter(WorkPlan.id == wp_id).first()
    if not wp:
        raise HTTPException(status_code=404, detail="Work plan not found")
    codes = (
        db.query(BudgetCode)
        .options(joinedload(BudgetCode.department))
        .filter(BudgetCode.work_plan_id == wp_id)
        .order_by(BudgetCode.id.asc())
        .all()
    )
    sources = db.query(RevenueSource).filter(RevenueSource.work_plan_id == wp_id).order_by(RevenueSource.id.asc()).all()
    committed_map = _bulk_committed_amounts(db, [c.id for c in codes])
    buf = build_workplan_report_pdf(wp, codes, committed_map, sources)
    fname = f"Annual_Work_Plan_FY_{wp.financial_year.replace('/', '-')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------- Dashboard & Reports ----------------------------

@app.get("/api/dashboard/stats")
def dashboard_stats(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    # Personal/role-scoped counters are cheap and always fresh (they depend
    # on who's asking), so they're computed directly. The heavier,
    # role-independent part of this payload — total budget, utilisation,
    # and the per-department chart data — is cached, since it is identical
    # for every viewer and was previously the slow part of this endpoint
    # (it used to call BudgetCode.committed_amount once per row, each of
    # which opened its own DB session — an N+1 pattern fixed below by
    # _bulk_committed_amounts).
    base = db.query(Requisition)
    if user.role == "staff":
        base = base.filter(Requisition.requester_id == user.id)
    elif user.role == "hod":
        base = base.filter(Requisition.department_id == user.department_id)

    pending = base.filter(Requisition.current_stage.in_(["hod", "treasurer", "clerk"])).count()
    approved = base.filter(Requisition.status.in_(["approved", "accounted"])).count()
    rejected = base.filter(Requisition.status == "rejected").count()
    recent = base.order_by(Requisition.created_at.desc()).limit(6).all()

    budget_summary = _cache_get("dashboard_stats:budget_summary")
    if budget_summary is None:
        # The dashboard's Total Budget figure is the total of the workplan
        # budget across every Annual Work Plan on record — it is not tied
        # to any one financial year, so it isn't scoped to the current
        # (most recently created) work plan the way the Work Plan & Budget
        # screen's GRAND TOTAL is.
        codes_q = db.query(BudgetCode).options(joinedload(BudgetCode.department))
        all_codes = codes_q.all()
        committed_map = _bulk_committed_amounts(db, [bc.id for bc in all_codes])

        total_budget_sum = sum(bc.allocated_amount for bc in all_codes)
        utilized = sum(committed_map.get(bc.id, 0.0) for bc in all_codes)

        dept_totals: dict = {}
        for bc in all_codes:
            dept_name = bc.department.name if bc.department else "Unassigned"
            dept_totals[dept_name] = dept_totals.get(dept_name, 0.0) + bc.allocated_amount
        budget_by_department = sorted(
            [{"department": name, "amount": amount} for name, amount in dept_totals.items()],
            key=lambda d: d["amount"], reverse=True,
        )

        budget_summary = {
            "total_budget": total_budget_sum,
            "budget_utilized": utilized,
            "utilization_pct": round((utilized / total_budget_sum * 100), 1) if total_budget_sum else 0,
            "budget_by_department": budget_by_department,
        }
        _cache_set("dashboard_stats:budget_summary", budget_summary)

    return {
        "pending_approvals": pending,
        "approved_requisitions": approved,
        "rejected_requisitions": rejected,
        "total_budget": budget_summary["total_budget"],
        "budget_utilized": budget_summary["budget_utilized"],
        "utilization_pct": budget_summary["utilization_pct"],
        "budget_by_department": budget_summary["budget_by_department"],
        "recent_activity": [
            {"ref_no": r.ref_no, "status": r.status, "amount": r.amount_requested,
             "department": r.department.name if r.department else None,
             "created_at": r.created_at.isoformat()}
            for r in recent
        ],
    }


@app.get("/api/reports/audit-view/{req_id}")
def audit_view(req_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(Requisition).filter(Requisition.id == req_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Requisition not found")
    return requisition_to_dict(r)


@app.get("/api/reports/audit-logs")
def get_audit_logs(db: Session = Depends(get_db), user: User = Depends(require_roles("admin", "auditor", "clerk"))):
    logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(300).all()
    return [
        {"id": l.id, "user": l.user_id, "action": l.action, "details": l.details,
         "created_at": l.created_at.isoformat()}
        for l in logs
    ]


@app.get("/api/health")
def health():
    return {"status": "ok", "time": dt.datetime.utcnow().isoformat()}
